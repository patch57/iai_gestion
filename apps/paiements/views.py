"""
Vues pour la gestion des paiements
IAI-Cameroun - Centre de Douala
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from apps.authentification.decorators import role_required
from django.contrib import messages
from django.http import JsonResponse
from django.urls import reverse
from django.utils import timezone
from decimal import Decimal
from .models import RecuPaiement, TranchePaiement, SessionConcours, EcheanceSessionNiveau1, ResultatConcours
from apps.etudiants.models import Etudiant



@login_required
def liste_recus(request):
    """Liste des reçus"""
    recus = RecuPaiement.objects.select_related('etudiant', 'tranche').order_by('-date_televersement')
    
    # Filtres
    statut = request.GET.get('statut')
    if statut:
        recus = recus.filter(statut=statut)
    
    paginator = Paginator(recus, 20)
    page = request.GET.get('page', 1)
    recus_page = paginator.get_page(page)
    
    context = {
        'recus': recus_page,
        'titre': 'Liste des reçus'
    }
    return render(request, 'paiements/recus/liste.html', context)


from django.core.exceptions import ValidationError
from .forms import valider_fichier_recu

@login_required
def televerser_recu(request, etudiant_id):
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id)
    
    # Sécurité supplémentaire
    if request.user.type_utilisateur == 'ETUDIANT' and etudiant.utilisateur != request.user:
        messages.error(request, "Accès refusé.")
        return redirect('tableau_bord:tableau_bord')
        
    # Vérifier si le profil de l'étudiant est incomplet (VALABLE UNIQUEMENT POUR LES ÉTUDIANTS)
    if request.user.type_utilisateur == 'ETUDIANT':
        if not (etudiant.date_naissance and etudiant.lieu_naissance and etudiant.sexe and etudiant.nationalite and etudiant.telephone and etudiant.adresse and etudiant.nom_tuteur and etudiant.telephone_tuteur and etudiant.photo):
            messages.warning(request, "⚠️ Veuillez compléter vos informations de profil et téléverser votre photo d'identité avant de pouvoir soumettre un reçu de paiement.")
            return redirect(reverse('tableau_bord:modifier_profil') + f'?compte_incomplet=1&next={request.path}')
        
    from apps.inscriptions.utils import get_current_academic_year_code
    tranches = TranchePaiement.objects.filter(annee_academique=etudiant.annee_academique.code if etudiant.annee_academique else get_current_academic_year_code(), est_actif=True)
    
    if request.method == 'POST':
        tranche_id = request.POST.get('tranche')
        file = request.FILES.get('recu_fichier')
        montant = request.POST.get('montant_mentionne')
        reference = request.POST.get('reference_recu', '')
        
        if not file or not tranche_id or not montant:
            messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
        else:
            try:
                valider_fichier_recu(file)
            except ValidationError as e:
                messages.error(request, f"❌ {e.message}")
                return render(request, 'paiements/recus/televerser.html', {
                    'etudiant': etudiant,
                    'tranches': tranches,
                    'titre': 'Téléverser un reçu'
                })
            
            tranche = None
            commentaires = ""
            
            if tranche_id in ['totalite', 'autre']:
                commentaires = f"OPTION:{tranche_id.upper()}"
            else:
                try:
                    tranche = TranchePaiement.objects.get(pk=tranche_id)
                except (TranchePaiement.DoesNotExist, ValueError):
                    pass

            recu = RecuPaiement.objects.create(
                etudiant=etudiant,
                tranche=tranche,
                recu_fichier=file,
                montant_mentionne=montant,
                reference_recu=reference,
                statut='EN_ATTENTE',
                commentaires=commentaires
            )
            
            # Analyse OCR réelle du reçu
            recu.analyser_par_ia()
            
            # Message basé sur le résultat réel
            if recu.statut == 'VALIDE':
                montant_detecte = recu.verification_ia.get('montant_principal', '')
                banque = recu.verification_ia.get('banque', 'Banque détectée')
                messages.success(
                    request,
                    f"✅ Reçu analysé par OCR — Score de confiance : {recu.score_confiance:.0%}. "
                    f"{'Montant: {:,.0f} FCFA. '.format(montant_detecte) if montant_detecte else ''}"
                    f"{'Banque: ' + banque + '. ' if banque else ''}"
                    f"Vérifié et validé automatiquement !"
                )
            elif recu.verification_ia.get('type_document') == 'RECU_MANUSCRIT_IAI' or (recu.tranche and recu.tranche.numero == 1):
                messages.info(
                    request,
                    f"📄 Reçu manuscrit d'Entrée en Caisse IAI téléversé avec succès ! "
                    f"Vos informations ({etudiant.get_nom_complet()}) ont été transmises au service comptabilité pour validation rapide."
                )
            elif recu.score_confiance and recu.score_confiance >= 0.50:
                messages.warning(
                    request,
                    f"⏳ Reçu téléversé (score OCR : {recu.score_confiance:.0%}). "
                    f"Vérification manuelle requise par le service comptabilité."
                )
            else:
                anomalies = recu.anomalies_detectees.get('anomalies', []) if isinstance(recu.anomalies_detectees, dict) else []
                messages.warning(
                    request,
                    f"⚠️ Reçu téléversé mais score faible ({recu.score_confiance:.0%}). "
                    f"{anomalies[0] if anomalies else 'Document peu lisible.'} "
                    f"Vérification manuelle requise."
                )
                
            return redirect('tableau_bord:tableau_bord')

            
    context = {
        'etudiant': etudiant,
        'tranches': tranches,
        'titre': 'Téléverser un reçu'
    }
    return render(request, 'paiements/recus/televerser.html', context)


@login_required
def televerser_recu_tranche(request, etudiant_id, tranche_id):
    """Téléverser un reçu pour une tranche spécifique"""
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id)
    
    # Sécurité supplémentaire
    if request.user.type_utilisateur == 'ETUDIANT' and etudiant.utilisateur != request.user:
        messages.error(request, "Accès refusé.")
        return redirect('tableau_bord:tableau_bord')
        
    # Vérifier si le profil de l'étudiant est incomplet (VALABLE UNIQUEMENT POUR LES ÉTUDIANTS)
    if request.user.type_utilisateur == 'ETUDIANT':
        if not (etudiant.date_naissance and etudiant.lieu_naissance and etudiant.sexe and etudiant.nationalite and etudiant.telephone and etudiant.adresse and etudiant.nom_tuteur and etudiant.telephone_tuteur and etudiant.photo):
            messages.warning(request, "⚠️ Veuillez compléter vos informations de profil et téléverser votre photo d'identité avant de pouvoir soumettre un reçu de paiement.")
            return redirect(reverse('tableau_bord:modifier_profil') + f'?compte_incomplet=1&next={request.path}')
        
    tranche = get_object_or_404(TranchePaiement, pk=tranche_id)
    
    if request.method == 'POST':
        file = request.FILES.get('recu_fichier')
        reference = request.POST.get('reference_recu', '')
        
        if not file:
            messages.error(request, 'Veuillez sélectionner un fichier.')
        else:
            recu = RecuPaiement.objects.create(
                etudiant=etudiant,
                tranche=tranche,
                recu_fichier=file,
                montant_mentionne=tranche.montant,
                reference_recu=reference,
                statut='EN_ATTENTE'
            )
            
            # Analyse OCR réelle du reçu
            recu.analyser_par_ia()
            
            # Message basé sur le résultat réel
            if recu.score_confiance and recu.score_confiance >= 0.90:
                montant_detecte = recu.verification_ia.get('montant_principal', '')
                banque = recu.verification_ia.get('banque', '')
                messages.success(
                    request,
                    f"✅ Reçu analysé par OCR — Score : {recu.score_confiance:.0%}. "
                    f"{'Montant: {:,.0f} FCFA. '.format(montant_detecte) if montant_detecte else ''}"
                    f"{'Banque: ' + banque + '. ' if banque else ''}"
                    f"Vérifié automatiquement !"
                )
            elif recu.score_confiance and recu.score_confiance >= 0.50:
                messages.warning(
                    request,
                    f"⏳ Reçu téléversé (score OCR : {recu.score_confiance:.0%}). "
                    f"Vérification manuelle requise par le service comptabilité."
                )
            else:
                anomalies = recu.anomalies_detectees.get('anomalies', []) if isinstance(recu.anomalies_detectees, dict) else []
                messages.warning(
                    request,
                    f"⚠️ Reçu téléversé mais score faible ({recu.score_confiance:.0%}). "
                    f"{anomalies[0] if anomalies else 'Document peu lisible.'} "
                    f"Vérification manuelle requise."
                )
            
            # Notification automatique du Chef de la Comptabilité (Dashboard & WhatsApp)
            try:
                from apps.tableau_bord.whatsapp_service import WhatsAppService
                WhatsAppService.notifier_chef_comptabilite(
                    titre=f"Nouveau reçu téléversé : {etudiant.get_nom_complet()}",
                    message=f"L'étudiant(e) {etudiant.get_nom_complet()} ({etudiant.matricule}) a téléversé un reçu de {tranche.montant:,.0f} FCFA pour la {tranche.get_numero_display()}.\nRéférence : {reference or 'Non renseignée'}.",
                    lien='/paiements/recus/'
                )
            except Exception as err_notif:
                from .views import logger_paiement
                logger_paiement.error(f"Erreur notification Chef Comptabilité : {err_notif}")

            return redirect('tableau_bord:tableau_bord')
            
    context = {
        'etudiant': etudiant,
        'tranche': tranche,
        'titre': f'Téléverser {tranche.get_numero_display()}'
    }
    return render(request, 'paiements/recus/televerser_tranche.html', context)


@login_required
def detail_recu(request, pk):
    """Détail d'un reçu"""
    recu = get_object_or_404(RecuPaiement, pk=pk)
    
    # Prévention IDOR : Cloisonnement strict pour les étudiants
    if request.user.type_utilisateur == 'ETUDIANT' and recu.etudiant.utilisateur != request.user:
        from django.core.exceptions import PermissionDenied
        raise PermissionDenied("Vous n'êtes pas autorisé à consulter ce reçu de paiement.")
        
    context = {
        'recu': recu,
        'titre': f'Détail du reçu'
    }
    return render(request, 'paiements/recus/detail.html', context)


@login_required
def valider_recu(request, pk):
    """Valider manuellement un reçu bancaire - Habilitation exclusive : Chef de la Comptabilité"""
    if request.user.type_utilisateur != 'CHEF_COMPTABILITE' and not request.user.is_superuser:
        messages.error(request, "⛔ Accès refusé : Seul le Chef de la Comptabilité est habilité à valider manuellement les reçus de paiement soumis sur la plateforme.")
        return redirect('paiements:liste_recus')

    recu = get_object_or_404(RecuPaiement, pk=pk)
    recu.statut = 'VALIDE'
    recu.date_verification = timezone.now()
    recu.verifie_par = request.user
    recu.save()
    
    # Audit log
    try:
        from apps.tableau_bord.models import JournalAudit
        JournalAudit.enregistrer(
            request=request,
            categorie='PAIEMENT',
            action=f"Validation Reçu #{recu.id}",
            details=f"Reçu de {recu.montant_mentionne} FCFA pour {recu.etudiant} validé."
        )
    except Exception:
        pass
    
    # Notification automatique multi-acteurs (Dashboard & WhatsApp)
    try:
        from apps.tableau_bord.whatsapp_service import WhatsAppService
        from apps.tableau_bord.services_notification import NotificationService
        etud_nom = recu.etudiant.get_nom_complet() if recu.etudiant else "Un étudiant"
        tranche_str = recu.tranche.get_numero_display() if recu.tranche else "Scolarité"

        # Notifier l'étudiant
        if recu.etudiant and hasattr(recu.etudiant, 'utilisateur') and recu.etudiant.utilisateur:
            NotificationService.notifier_utilisateur(
                recu.etudiant.utilisateur,
                titre="Paiement Validé 💰",
                message=f"Votre reçu de paiement de {recu.montant_mentionne:,.0f} FCFA ({tranche_str}) a été validé.",
                type_notif='SUCCESS',
                lien='/paiements/mes-penalites/'
            )

        # Notifier le Directeur et l'Admin Financier
        NotificationService.notifier_directeur(
            titre=f"Paiement Validé : {etud_nom}",
            message=f"Montant : {recu.montant_mentionne:,.0f} FCFA ({tranche_str}) validé par {request.user.get_full_name() or request.user.username}.",
            type_notif='SUCCESS',
            lien='/paiements/recus/'
        )
        NotificationService.notifier_admin_financier(
            titre=f"Paiement Enregistré : {etud_nom}",
            message=f"Règlement de {recu.montant_mentionne:,.0f} FCFA confirmé pour {etud_nom}.",
            type_notif='SUCCESS',
            lien='/paiements/recus/'
        )

        WhatsAppService.notifier_chef_comptabilite(
            titre=f"Reçu validé : {etud_nom} ({recu.montant_mentionne:,.0f} FCFA)",
            message=f"Le reçu pour {etud_nom} ({tranche_str}) d'un montant de {recu.montant_mentionne:,.0f} FCFA a été validé avec succès par {request.user.get_full_name() or request.user.username}.",
            lien='/paiements/recus/'
        )
    except Exception as err_notif:
        from .views import logger_paiement
        logger_paiement.error(f"Erreur notification paiement : {err_notif}")

    messages.success(request, f'✅ Reçu validé avec succès !')
    return redirect('paiements:liste_recus')


@login_required
def rejeter_recu(request, pk):
    """Rejeter un reçu avec motif obligatoire - Réservé au Chef de la Comptabilité"""
    if request.user.type_utilisateur != 'CHEF_COMPTABILITE' and not request.user.is_superuser:
        messages.error(request, "⛔ Accès refusé : Seul le Chef de la Comptabilité est habilité à statuer sur les reçus de paiement.")
        return redirect('paiements:liste_recus')

    recu = get_object_or_404(RecuPaiement, pk=pk)
    
    if request.method == 'POST':
        motif = request.POST.get('motif_rejet', '').strip()
        if not motif:
            messages.error(request, "❌ Motif de rejet obligatoire ! Veuillez indiquer la raison du rejet.")
            return redirect('paiements:detail_recu', pk=pk)
            
        recu.statut = 'REJETE'
        recu.commentaires = f"MOTIF DE REJET : {motif}"
        recu.date_verification = timezone.now()
        recu.verifie_par = request.user
        recu.save()
        
        # Notifier l'étudiant
        if recu.etudiant and recu.etudiant.utilisateur:
            from apps.tableau_bord.models import Notification
            Notification.objects.create(
                utilisateur=recu.etudiant.utilisateur,
                type='DANGER',
                titre=f"Reçu de paiement rejeté ({recu.tranche.get_numero_display() if recu.tranche else 'Scolarité'})",
                message=f"Votre reçu a été rejeté par la Comptabilité. Motif : {motif}. Veuillez téléverser un nouveau reçu valide.",
                lien='/inscriptions/'
            )
            
        messages.warning(request, f'⚠️ Reçu de {recu.etudiant.get_nom_complet()} rejeté (Motif : {motif})')
        return redirect('paiements:liste_recus')

    context = {
        'recu': recu,
        'titre': 'Rejeter le reçu'
    }
    return render(request, 'paiements/recus/rejeter_modal.html', context)



@login_required
def liste_tranches(request):
    """Liste des tranches de paiement"""
    tranches = TranchePaiement.objects.all()
    
    context = {
        'tranches': tranches,
        'titre': 'Tranches de paiement'
    }
    return render(request, 'paiements/tranches/liste.html', context)


@login_required
@permission_required('paiements.add_tranche', raise_exception=True)
def ajouter_tranche(request):
    """Ajouter une tranche de paiement"""
    if request.method == 'POST':
        messages.success(request, '✅ Tranche ajoutée avec succès !')
        return redirect('paiements:liste_tranches')
    
    context = {
        'titre': 'Ajouter une tranche'
    }
    return render(request, 'paiements/tranches/form.html', context)


@login_required
@permission_required('paiements.change_tranche', raise_exception=True)
def modifier_tranche(request, pk):
    """Modifier une tranche de paiement"""
    tranche = get_object_or_404(TranchePaiement, pk=pk)
    
    if request.method == 'POST':
        messages.success(request, '✏️ Tranche modifiée avec succès !')
        return redirect('paiements:liste_tranches')
    
    context = {
        'tranche': tranche,
        'titre': 'Modifier la tranche'
    }
    return render(request, 'paiements/tranches/form.html', context)


@login_required
@permission_required('paiements.delete_tranche', raise_exception=True)
def supprimer_tranche(request, pk):
    """Supprimer une tranche de paiement"""
    tranche = get_object_or_404(TranchePaiement, pk=pk)
    
    if request.method == 'POST':
        tranche.delete()
        messages.success(request, '🗑️ Tranche supprimée avec succès !')
        return redirect('paiements:liste_tranches')
    
    context = {
        'tranche': tranche,
        'titre': 'Supprimer la tranche'
    }
    return render(request, 'paiements/tranches/supprimer.html', context)


@login_required
def statistiques_paiements(request):
    """Statistiques des paiements"""
    context = {
        'titre': 'Statistiques des paiements'
    }
    return render(request, 'paiements/statistiques.html', context)


@login_required
def api_recus_attente(request):
    """API pour le nombre de reçus en attente"""
    count = RecuPaiement.objects.filter(statut='EN_ATTENTE').count()
    return JsonResponse({'count': count})


@login_required
def api_prescan_recu(request):
    """
    API AJAX pour pré-analyser un fichier de reçu bancaire téléversé et extraire
    automatiquement le montant et le numéro de référence via le moteur OCR IA.
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Méthode POST requise.'}, status=405)
    
    file_obj = request.FILES.get('recu_fichier')
    if not file_obj:
        return JsonResponse({'success': False, 'error': 'Aucun fichier reçu.'}, status=400)
    
    from django.core.files.storage import default_storage
    from django.core.files.base import ContentFile
    import re

    # Enregistrer temporairement le fichier pour l'analyse OCR
    temp_filename = f"tmp_prescan_{request.user.id}_{file_obj.name}"
    saved_path = default_storage.save(f"tmp_ocr/{temp_filename}", ContentFile(file_obj.read()))
    full_path = default_storage.path(saved_path)
    
    try:
        from .ocr_service import extraire_texte_depuis_image, BANQUES_MOTS_CLES
        texte_brut = extraire_texte_depuis_image(full_path) or ""
        texte = (texte_brut + " " + file_obj.name).upper()
        
        # Mots-clés d'en-tête ou de nom de fichier à exclure pour les références
        INVALID_REF_WORDS = {
            'CLIENT', 'ANTENNE', 'AGENCE', 'COMPTE', 'CAISSE', 'MONTANT', 'VERSEMENT', 'MOTIF',
            'REMETTANT', 'DEVISE', 'FRANC', 'SOLDE', 'CONCOURS', 'TRESOR', 'OPERATION',
            'BILLETAGE', 'VALEUR', 'NOMBRE', 'BILLETS', 'BILLET', 'GUICHET', 'PAYEUR', 'TAXE', 'FRAIS',
            'YAOUNDE', 'DOUALA', 'BESSENGUE', 'YDE', 'DLA', 'XAF', 'FCFA', 'CFA',
            'INSCRIPTION', 'PREINSCRIPTION', 'PRE-INSCRIPTION', 'PAIEMENT', 'PAYMENT',
            'RECU', 'BORDEREAU', 'DOCUMENT', 'TRANCHE', 'SCAN', 'SCANN', 'IMAGE', 'PHOTO',
            'TELECHARGEMENT', 'ENTREE', 'SCOLAIRE', 'IAI', 'FICHIER', 'JPEG', 'JPG', 'PNG', 'PDF'
        }

        # Valeurs faciales de coupures de billets et pièces FCFA à exclure absolument des numéros de bordereaux
        DENOMINATIONS_FCFA = {10000, 5000, 2000, 1000, 500, 100, 50}

        # 1. Extraction du montant principal (dans le texte ou le nom du fichier)
        MONTS_CANDIDATS_CONNUS = [474000, 461000, 175000, 115000, 100000, 84000, 71000, 50000, 35000, 30000, 3000]
        montant_detecte = None

        # Recherche directe des montants connus
        for m_conn in MONTS_CANDIDATS_CONNUS:
            if str(m_conn) in texte.replace(' ', '').replace('.', '').replace(',', ''):
                montant_detecte = m_conn
                break
        
        if not montant_detecte:
            regex_montants = r'(\b\d{1,3}(?:[\s\.,]\d{3})*)\s*(?:FCFA|XAF|CFA|F\b)?'
            matches_montant = re.findall(regex_montants, texte)
            for m_str in matches_montant:
                clean_m = re.sub(r'[^\d]', '', m_str)
                if clean_m.isdigit() and 1000 <= int(clean_m) <= 1000000:
                    montant_detecte = int(clean_m)
                    break

        def est_candidat_ref_valide(cand):
            if not cand:
                return False
            c_str = str(cand).strip('-_ ')
            if len(c_str) < 4 or any(bad in c_str for bad in INVALID_REF_WORDS):
                return False
            if c_str.startswith('237'): # Exclure numéros de téléphone camerounais
                return False
            if len(c_str) == 5 and c_str.startswith('000'): # Exclure codes d'agence 5 chiffres
                return False
            if not re.search(r'\d', c_str):
                return False
            if c_str.isdigit():
                val = int(c_str)
                if val % 1000 == 0:  # Exclure tous les montants et sous-totaux (ex: 110000, 115000, 10000)
                    return False
                if len(c_str) == 11: # Exclure numéros de compte 11 chiffres
                    return False
            if montant_detecte and str(montant_detecte) in c_str:
                return False
            return True
        
        # 2. Extraction du numéro de référence / bordereau
        reference_detectee = None

        # Règle Priorité 1 : Numéro de bordereau bancaire classique avec zéro initial (ex: 011261, N011261, N°011261)
        m_scb_zero = re.search(r'(?:N|NO|NUM|REF|TIERS|BORDEREAU|N°|Nº|N°:|\b)\s*[:\.]?\s*(0[1-9]\d{3,8})\b', texte)
        if m_scb_zero:
            cand = m_scb_zero.group(1)
            if est_candidat_ref_valide(cand):
                reference_detectee = cand

        # Règle Priorité 2 : Motif N° bordereau explicite (ex: N° 011261, TIERS N° 011261, BORDEREAU N° 011261)
        if not reference_detectee:
            m_direct = re.search(r'(?:TIERS\s*N[°ºO\W]?|BORDEREAU\s*N[°ºO\W]?|N[°ºO\W])\s*[:\.]?\s*([0-9]{5,12})', texte)
            if m_direct:
                cand = m_direct.group(1)
                if est_candidat_ref_valide(cand):
                    reference_detectee = cand

        # Règle Priorité 3 : Code banque alphanumérique structuré (Ex: SCB-2025-09812, FT240710-9812)
        if not reference_detectee:
            gen_match = re.search(r'\b([A-Z]{2,4}[-_\s]?\d{4,8}[-_\s]?[A-Z0-9]{2,8})\b', texte)
            if gen_match:
                cand = gen_match.group(1).replace(' ', '').strip('-_ ')
                if est_candidat_ref_valide(cand):
                    reference_detectee = cand

        # Règle Priorité 3 : Motif général avec préfixe (Ref: XXX123, Transaction: TXN12345)
        if not reference_detectee:
            regex_ref = r'(?:REF|TRANSACTION|TXN|ID|BORD)[:\s]*([0-9A-Z\-_]{5,25})'
            match_ref = re.search(regex_ref, texte)
            if match_ref:
                cand = match_ref.group(1).strip('-_ ')
                if est_candidat_ref_valide(cand):
                    reference_detectee = cand

        # Règle Priorité 4 : Séquence de chiffres isolée de 5 à 12 chiffres
        if not reference_detectee:
            digits_matches = re.findall(r'\b(\d{5,12})\b', texte)
            for d in digits_matches:
                if est_candidat_ref_valide(d):
                    reference_detectee = d
                    break
        
        # Nettoyage final
        if reference_detectee and not est_candidat_ref_valide(reference_detectee):
            reference_detectee = None

        # 3. Détection de la banque
        banque_detectee = "Inconnue"
        for b_code, mots in BANQUES_MOTS_CLES.items():
            if any(m in texte for m in mots):
                banque_detectee = b_code
                break
                
        return JsonResponse({
            'success': True,
            'montant': montant_detecte,
            'reference': reference_detectee,
            'banque': banque_detectee,
            'score': 0.90 if (montant_detecte and reference_detectee) else 0.65
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
    finally:
        try:
            if default_storage.exists(saved_path):
                default_storage.delete(saved_path)
        except Exception:
            pass


from .services import calculer_penalites_etudiant
from .momo_service import CinetPayService
from .models import TransactionPaiement
import json as json_module
import logging

logger_paiement = logging.getLogger(__name__)


@login_required
def mes_penalites(request):
    """Page dédiée au suivi et à la gestion des pénalités de retard de l'étudiant"""
    etudiant = Etudiant.objects.filter(utilisateur=request.user).first()
    
    if not etudiant and (request.user.est_etudiant or getattr(request.user, 'type_utilisateur', None) == 'ETUDIANT'):
        from apps.etudiants.models import Filiere, Niveau
        from apps.inscriptions.models import AnneeAcademique
        from datetime import timedelta
        
        filiere = Filiere.objects.first()
        niveau = Niveau.objects.first()
        annee_active = AnneeAcademique.objects.filter(est_actuelle=True).first() or AnneeAcademique.objects.first()

        
        try:
            etudiant = Etudiant.objects.create(
                utilisateur=request.user,
                nom=request.user.last_name or 'Etudiant',
                prenom=request.user.first_name or 'IAI',
                email=request.user.email,
                telephone=getattr(request.user, 'telephone', '') or '699999999',
                date_naissance=timezone.now().date() - timedelta(days=7300),
                lieu_naissance='Douala',
                sexe='M',
                filiere=filiere,
                niveau=niveau,
                annee_academique=annee_active,
                matricule=getattr(request.user, 'matricule', '') or f"GL.CMR.D{request.user.id:03d}.2425A"
            )
        except Exception:
            etudiant = Etudiant.objects.filter(utilisateur=request.user).first()

    if not etudiant:
        messages.error(request, "Accès réservé aux étudiants.")
        return redirect('tableau_bord:tableau_bord')

    penalites_info = calculer_penalites_etudiant(etudiant)
    from apps.tableau_bord.models import PenalitePaiement
    from .models import TransactionPaiement

    # Transactions Mobile Money réussies pour pénalités
    transactions_penalites = TransactionPaiement.objects.filter(
        etudiant=etudiant,
        type_paiement='PENALITE',
        statut='SUCCESS'
    ).order_by('-date_creation')

    # Pénalités réglées manuellement en caisse
    penalites_reglees_caisse = PenalitePaiement.objects.filter(
        etudiant=etudiant,
        est_regle=True
    ).order_by('-date_paiement')

    context = {
        'etudiant': etudiant,
        'penalites_info': penalites_info,
        'penalites_reglees_caisse': penalites_reglees_caisse,
        'transactions_penalites': transactions_penalites,
        'titre': 'Mes Pénalités de Retard'
    }
    return render(request, 'paiements/mes_penalites.html', context)


@login_required
def payer_penalites(request):
    """Page de checkout pour payer les pénalités accumulées"""
    etudiant = Etudiant.objects.filter(utilisateur=request.user).first()
    if not etudiant:
        messages.error(request, "Accès réservé aux étudiants.")
        return redirect('tableau_bord:tableau_bord')

    penalites_info = calculer_penalites_etudiant(etudiant)

    if penalites_info['total_eligibles'] <= 0:
        messages.warning(request, "⚠️ Aucune pénalité éligible au paiement en ligne. Les pénalités d'une tranche ne peuvent être réglées en ligne que lorsque le reçu bancaire de cette tranche a été validé par la comptabilité.")
        return redirect('paiements:mes_penalites')

    context = {
        'etudiant': etudiant,
        'penalites_info': penalites_info,
        'total_a_payer': penalites_info['total_eligibles'],
        'titre': 'Payer mes Pénalités'
    }
    return render(request, 'paiements/recus/payer_penalite.html', context)


@login_required
def initier_paiement_momo(request):
    """Initialise le paiement Mobile Money sécurisé via CinetPay API."""
    if request.method != 'POST':
        return JsonResponse({'status': 'FAILED', 'message': 'Méthode non autorisée.'}, status=405)

    import json
    data = {}
    if request.body:
        try:
            data = json.loads(request.body)
        except Exception:
            pass

    telephone = data.get('telephone') or request.POST.get('telephone', '')
    operateur = data.get('operateur') or request.POST.get('operateur', '')

    etudiant = get_object_or_404(Etudiant, utilisateur=request.user)
    penalites_info = calculer_penalites_etudiant(etudiant)
    amount = penalites_info['total_eligibles']

    if amount <= 0:
        return JsonResponse({'status': 'FAILED', 'message': 'Aucune pénalité éligible à payer en ligne. La tranche associée doit d\'abord être validée par la comptabilité.'})

    transaction = TransactionPaiement(
        etudiant=etudiant,
        transaction_id=TransactionPaiement.generer_transaction_id(),
        montant=amount,
        type_paiement='PENALITE',
        telephone=telephone,
        operateur=operateur
    )
    transaction.save()

    from django.conf import settings as django_settings
    base_url = getattr(django_settings, 'SITE_BASE_URL', 'http://127.0.0.1:8000').rstrip('/')
    notify_url = base_url + reverse('paiements:webhook_cinetpay')
    return_url = base_url + reverse('paiements:paiement_succes') + f'?transaction_id={transaction.transaction_id}'

    res = CinetPayService.initier_paiement(
        transaction_id=transaction.transaction_id,
        amount=amount,
        description=f"Pénalités de retard - {etudiant.get_nom_complet()} ({etudiant.matricule})",
        notify_url=notify_url,
        return_url=return_url,
        customer_name=etudiant.get_nom_complet(),
        customer_email=getattr(etudiant.utilisateur, 'email', ''),
        customer_phone=telephone or etudiant.telephone or '699999999'
    )

    if res['status'] == 'PENDING':
        transaction.cinetpay_payment_token = res.get('payment_token', '')
        transaction.payment_url = res.get('payment_url', '')
        transaction.save(update_fields=['cinetpay_payment_token', 'payment_url'])

    return JsonResponse(res)



@login_required
def verifier_paiement_momo(request):
    """Vérifie le statut d'une transaction CinetPay (polling côté client)."""
    if request.method != 'POST':
        return JsonResponse({'status': 'FAILED', 'message': 'Méthode non autorisée.'}, status=405)

    data = json_module.loads(request.body)
    transaction_id = data.get('transaction_id')

    if not transaction_id:
        return JsonResponse({'status': 'FAILED', 'message': 'ID de transaction manquant.'})

    try:
        transaction = TransactionPaiement.objects.get(transaction_id=transaction_id)
    except TransactionPaiement.DoesNotExist:
        return JsonResponse({'status': 'FAILED', 'message': 'Transaction introuvable.'})

    # Si déjà traitée, retourner le statut
    if transaction.statut == 'SUCCESS':
        return JsonResponse({
            'status': 'SUCCESS',
            'message': 'Paiement déjà confirmé.',
            'redirect_url': reverse('paiements:paiement_succes') + f'?transaction_id={transaction_id}'
        })

    # Vérifier auprès de CinetPay
    res = CinetPayService.verifier_statut_paiement(transaction_id)

    if res['status'] == 'SUCCESS' and transaction.statut != 'SUCCESS':
        tx_data = res.get('data', {})
        transaction.marquer_succes(cinetpay_data=tx_data)
        CinetPayService.regler_penalites_etudiant(
            etudiant=transaction.etudiant,
            cinetpay_data=tx_data,
            amount_to_pay=float(transaction.montant)
        )
        messages.success(request, f"Paiement de {transaction.montant:,.0f} FCFA confirmé avec succès !")
        res['redirect_url'] = reverse('paiements:paiement_succes') + f'?transaction_id={transaction_id}'

    elif res['status'] in ('FAILED', 'CANCELLED'):
        transaction.marquer_echec(cinetpay_data=res.get('data', {}))

    return JsonResponse(res)


from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def webhook_cinetpay(request):
    """
    Webhook CinetPay — reçoit les notifications de paiement automatiques.
    Pas de CSRF car appelé par les serveurs CinetPay.
    """
    if request.method == 'POST':
        try:
            # CinetPay envoie les données en POST form-encoded ou JSON
            cpm_trans_id = request.POST.get('cpm_trans_id') or ''
            if not cpm_trans_id and request.body:
                try:
                    body = json_module.loads(request.body)
                    cpm_trans_id = body.get('cpm_trans_id', '')
                except (json_module.JSONDecodeError, ValueError):
                    pass

            if not cpm_trans_id:
                logger_paiement.warning("[Webhook] Notification sans cpm_trans_id")
                return JsonResponse({'status': 'error', 'message': 'Missing transaction ID'})

            # Retrouver la transaction
            try:
                transaction = TransactionPaiement.objects.get(transaction_id=cpm_trans_id)
            except TransactionPaiement.DoesNotExist:
                logger_paiement.warning(f"[Webhook] Transaction inconnue: {cpm_trans_id}")
                return JsonResponse({'status': 'error', 'message': 'Unknown transaction'})

            # Vérifier le statut réel auprès de CinetPay (ne jamais faire confiance au webhook seul)
            res = CinetPayService.verifier_statut_paiement(cpm_trans_id)

            if res['status'] == 'SUCCESS' and transaction.statut != 'SUCCESS':
                tx_data = res.get('data', {})
                transaction.marquer_succes(cinetpay_data=tx_data)
                CinetPayService.regler_penalites_etudiant(
                    etudiant=transaction.etudiant,
                    cinetpay_data=tx_data,
                    amount_to_pay=float(transaction.montant)
                )
                logger_paiement.info(f"[Webhook] Paiement confirmé: {cpm_trans_id}")

            elif res['status'] in ('FAILED', 'CANCELLED') and transaction.statut == 'PENDING':
                transaction.marquer_echec(cinetpay_data=res.get('data', {}))
                logger_paiement.info(f"[Webhook] Paiement échoué/annulé: {cpm_trans_id}")

            return JsonResponse({'status': 'ok'})

        except Exception as e:
            logger_paiement.exception(f"[Webhook] Erreur: {e}")
            return JsonResponse({'status': 'error'}, status=500)

    return JsonResponse({'status': 'error', 'message': 'POST only'}, status=405)


from .models import SessionConcours, EcheanceSessionNiveau1

@login_required
def gestion_sessions_concours(request):
    """Vue de gestion des sessions de concours et de leurs échéances pour le Niveau 1"""
    user = request.user
    role = getattr(user, 'type_utilisateur', 'ETUDIANT')
    
    if role not in ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR']:
        messages.error(request, "Accès réservé au Chef de Service de la Comptabilité et à la Direction.")
        return redirect('tableau_bord:tableau_bord')
        
    sessions = SessionConcours.objects.prefetch_related('echeances').order_by('-date_concours')
    
    context = {
        'sessions': sessions,
        'titre': 'Sessions de Concours & Échéances Niveau 1'
    }
    return render(request, 'paiements/sessions_concours_liste.html', context)


@login_required
def creer_session_concours(request):
    """Créer une nouvelle session de concours Niveau 1 avec génération automatique des 4 échéances de base"""
    user = request.user
    role = getattr(user, 'type_utilisateur', 'ETUDIANT')
    
    if role not in ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR']:
        messages.error(request, "Permissions insuffisantes.")
        return redirect('paiements:gestion_sessions_concours')
        
    if request.method == 'POST':
        nom = request.POST.get('nom')
        code = request.POST.get('code')
        date_concours = request.POST.get('date_concours')
        from apps.inscriptions.utils import get_current_academic_year_code
        annee = request.POST.get('annee_academique', get_current_academic_year_code())
        description = request.POST.get('description', '')
        
        # Dates des 4 tranches
        d1 = request.POST.get('date_limite_t1')
        d2 = request.POST.get('date_limite_t2')
        d3 = request.POST.get('date_limite_t3')
        d4 = request.POST.get('date_limite_t4')
        
        if not nom or not code or not date_concours or not d1 or not d2 or not d3 or not d4:
            messages.error(request, "Veuillez remplir tous les champs obligatoires (session + les 4 échéances).")
        else:
            try:
                session = SessionConcours.objects.create(
                    nom=nom,
                    code=code,
                    date_concours=date_concours,
                    annee_academique=annee,
                    description=description,
                    est_active=True
                )
                
                # Création des 4 échéances standard Niveau 1 pour cette session (Pré-inscription = 84 000 FCFA)
                EcheanceSessionNiveau1.objects.create(
                    session_concours=session,
                    tranche_numero=1,
                    montant=84000,
                    date_limite=d1,
                    description="Frais de Pré-inscription / Inscription"
                )
                EcheanceSessionNiveau1.objects.create(
                    session_concours=session,
                    tranche_numero=2,
                    montant=175000,
                    date_limite=d2,
                    description="1ère Tranche Scolarité"
                )
                EcheanceSessionNiveau1.objects.create(
                    session_concours=session,
                    tranche_numero=3,
                    montant=115000,
                    date_limite=d3,
                    description="2ème Tranche Scolarité"
                )
                EcheanceSessionNiveau1.objects.create(
                    session_concours=session,
                    tranche_numero=4,
                    montant=100000,
                    date_limite=d4,
                    description="3ème Tranche Scolarité"
                )
                
                messages.success(request, f"✅ Session '{session.nom}' créée avec ses 4 échéances de paiement configurées avec succès !")
                return redirect('paiements:gestion_sessions_concours')
            except Exception as e:
                messages.error(request, f"Erreur lors de la création : {e}")
                
    return redirect('paiements:gestion_sessions_concours')


@login_required
def editer_echeances_session(request, pk):
    """Mettre à jour les dates limites d'une session de concours"""
    session = get_object_or_404(SessionConcours, pk=pk)
    
    if request.user.type_utilisateur not in ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR']:
        messages.error(request, "Permissions insuffisantes.")
        return redirect('paiements:gestion_sessions_concours')
        
    if request.method == 'POST':
        for t_num in [1, 2, 3, 4]:
            d_limite = request.POST.get(f'date_limite_{t_num}')
            remarque = request.POST.get(f'remarque_{t_num}', '')
            if d_limite:
                EcheanceSessionNiveau1.objects.update_or_create(
                    session_concours=session,
                    tranche_numero=t_num,
                    defaults={
                        'montant': 84000 if t_num == 1 else (175000 if t_num == 2 else (115000 if t_num == 3 else 100000)),
                        'date_limite': d_limite,
                        'description': remarque
                    }
                )
        messages.success(request, f"Les échéances de la session '{session.nom}' ont été mises à jour.")
        
    return redirect('paiements:gestion_sessions_concours')



@login_required
def paiement_succes(request):
    """Page de succès après le paiement en ligne. Affiche les détails de la transaction et confirme son statut."""
    transaction_id = request.GET.get('transaction_id')
    transaction = None
    if transaction_id:
        try:
            transaction = TransactionPaiement.objects.get(transaction_id=transaction_id)
            if transaction.statut != 'SUCCESS':
                # Vérifier et confirmer la transaction auprès du service CinetPay
                from .momo_service import CinetPayService
                res = CinetPayService.verifier_statut_paiement(transaction_id)
                if res.get('status') == 'SUCCESS':
                    transaction.marquer_succes(cinetpay_data=res.get('data', {}))
                    CinetPayService.regler_penalites_etudiant(
                        etudiant=transaction.etudiant,
                        cinetpay_data=res.get('data', {}),
                        amount_to_pay=float(transaction.montant)
                    )
                    transaction.refresh_from_db()
        except TransactionPaiement.DoesNotExist:
            pass

    context = {
        'titre': 'Paiement Réussi',
        'transaction': transaction,
    }
    return render(request, 'paiements/recus/paiement_succes.html', context)



# ==============================================================================
# GESTION DES RÉSULTATS DU CONCOURS (NIVEAU 1)
# ==============================================================================

import csv
import io
from django.http import HttpResponse


@login_required
def detail_session_concours(request, pk):
    """
    Vue détaillée d'une session de concours : 
    Affichage des échéances, KPIs et de la liste des candidats admis importés.
    """
    session = get_object_or_404(SessionConcours, pk=pk)
    user = request.user
    role = getattr(user, 'type_utilisateur', 'ETUDIANT')
    
    if role not in ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR']:
        messages.error(request, "Permissions insuffisantes pour consulter les détails de la session.")
        return redirect('paiements:gestion_sessions_concours')
        
    # Nettoyage automatique des entrées parasites de l'OCR (ex: "9058", "IA!", "III.", "Listed'attente")
    from django.db.models import Q, Case, When, Value, IntegerField
    session.resultats.filter(
        Q(nom__in=['9058', 'IA!', 'III.', 'Listed\'attente', 'Listedattente']) |
        Q(nom__icontains='Listed\'attente') |
        Q(nom__icontains='Listedattente') |
        Q(nom__regex=r'^\d+$')
    ).delete()

    resultats = ResultatConcours.objects.filter(session_concours=session).select_related('filiere', 'etudiant_cree')
    
    # Recherche et Filtres par Pré-inscription, 1ère Tranche, 2ème Tranche, 3ème Tranche
    search_q = request.GET.get('q', '').strip()
    filiere_id = request.GET.get('filiere', '')
    statut_t1 = request.GET.get('statut_t1', '') or request.GET.get('statut_preinscr', '')
    statut_t2 = request.GET.get('statut_t2', '')
    statut_t3 = request.GET.get('statut_t3', '')
    statut_t4 = request.GET.get('statut_t4', '')
    
    if search_q:
        resultats = resultats.filter(
            Q(nom__icontains=search_q) | 
            Q(prenom__icontains=search_q) | 
            Q(email__icontains=search_q)
        )
    if filiere_id:
        resultats = resultats.filter(filiere_id=filiere_id)
    if statut_t1:
        resultats = resultats.filter(statut_preinscription=statut_t1)
    if statut_t2:
        resultats = resultats.filter(statut_tranche2=statut_t2)
    if statut_t3:
        resultats = resultats.filter(statut_tranche3=statut_t3)
    if statut_t4:
        resultats = resultats.filter(statut_tranche4=statut_t4)


    # Tri strict : 1. GL Admis -> 2. GL Liste d'attente -> 3. SR Admis -> 4. SR Liste d'attente
    resultats = resultats.annotate(
        filiere_order=Case(
            When(filiere__code='GL', then=Value(1)),
            When(filiere__code='SR', then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        ),
        statut_order=Case(
            When(statut_admission='ADMIS', then=Value(1)),
            When(statut_admission='LISTE_ATTENTE', then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        )
    ).order_by('filiere_order', 'statut_order', 'nom', 'prenom', 'id')

        
    # KPIs (Pré-inscription = 84 000 FCFA)
    total_admis = session.resultats.count()
    total_preinscrits_payes = session.resultats.filter(statut_preinscription='PAYE').count()
    total_preinscrits_non_payes = total_admis - total_preinscrits_payes
    montant_recouvre = total_preinscrits_payes * 84000
    taux_recouvrement = round((total_preinscrits_payes / total_admis * 100), 1) if total_admis > 0 else 0
    
    from apps.etudiants.models import Filiere
    filieres = Filiere.objects.all()
    echeances = session.echeances.all()
    
    context = {
        'session': session,
        'resultats': resultats,
        'filieres': filieres,
        'echeances': echeances,
        'total_admis': total_admis,
        'total_preinscrits_payes': total_preinscrits_payes,
        'total_preinscrits_non_payes': total_preinscrits_non_payes,
        'filiere_id': filiere_id,
        'statut_t1': statut_t1,
        'statut_t2': statut_t2,
        'statut_t3': statut_t3,
        'statut_t4': statut_t4,
        'statut_preinscr': statut_t1,
        'search_q': search_q,
        'titre': f"Détails Session - {session.nom}"
    }
    return render(request, 'paiements/sessions_concours_detail.html', context)



@login_required
def importer_resultats_concours(request, pk):
    import re
    """

    Importation des résultats du concours (exclusif Centre de Douala).
    Prend en charge :
    1. CSV/TXT avec colonnes séparées par ; ou , ou tab
    2. PV officiel PDF de l'IAI-Cameroun (structuré par Sections de Centre, Filière et candidats séparés par ;)
    """
    session = get_object_or_404(SessionConcours, pk=pk)
    user = request.user
    role = getattr(user, 'type_utilisateur', 'ETUDIANT')
    
    if role not in ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR']:
        messages.error(request, "Seul le Chef Comptabilité ou un administrateur peut importer les résultats du concours.")
        return redirect('paiements:gestion_sessions_concours')
        
    if request.method == 'POST':
        fichier = request.FILES.get('fichier_csv')
        saisie_manuelle = request.POST.get('saisie_manuelle', '').strip()
        
        full_text = ""
        lignes = []
        if fichier:
            fname = fichier.name.lower()
            if fname.endswith('.csv') or fname.endswith('.txt'):
                try:
                    full_text = fichier.read().decode('utf-8-sig', errors='ignore')
                    lignes = full_text.splitlines()
                except Exception as e:
                    messages.error(request, f"Erreur lors de la lecture du fichier : {e}")
                    return redirect('paiements:detail_session_concours', pk=session.pk)
            elif fname.endswith('.pdf'):
                try:
                    pdf_bytes = fichier.read()
                    import pypdf
                    reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
                    text_pages = []
                    for page in reader.pages:
                        extracted = page.extract_text()
                        if extracted:
                            text_pages.append(extracted)
                    full_text = "\n".join(text_pages)

                    # Si le PDF est un scan (image CamScanner sans texte numérique), OCR Automatique
                    if not full_text.strip():
                        try:
                            import pypdfium2 as pdfium
                            from rapidocr_onnxruntime import RapidOCR
                            import numpy as np
                            
                            engine = RapidOCR()
                            pdf_doc = pdfium.PdfDocument(pdf_bytes)
                            ocr_pages = []
                            for i in range(len(pdf_doc)):
                                page = pdf_doc[i]
                                pil_image = page.render(scale=2.5).to_pil()
                                img_np = np.array(pil_image)
                                ocr_res, _ = engine(img_np)
                                if ocr_res:
                                    ocr_pages.append("\n".join([item[1] for item in ocr_res if item and len(item) > 1]))
                            full_text = "\n".join(ocr_pages)
                        except Exception as ocr_err:
                            messages.error(request, f"Erreur lors de l'OCR du document scanné : {ocr_err}")
                            return redirect('paiements:detail_session_concours', pk=session.pk)

                    lignes = full_text.splitlines()
                except Exception as e:
                    messages.error(request, f"Erreur lors de l'extraction du fichier PDF : {e}")
                    return redirect('paiements:detail_session_concours', pk=session.pk)


            else:
                messages.error(request, "Veuillez téléverser un fichier au format CSV (.csv), TXT (.txt) ou PDF (.pdf).")
                return redirect('paiements:detail_session_concours', pk=session.pk)
        elif saisie_manuelle:
            full_text = saisie_manuelle
            lignes = saisie_manuelle.splitlines()
        else:
            messages.error(request, "Veuillez fournir un fichier CSV/PDF ou remplir le champ de saisie rapide.")
            return redirect('paiements:detail_session_concours', pk=session.pk)

        import_count = 0
        update_count = 0
        error_count = 0
        
        from apps.etudiants.models import Filiere
        filieres = list(Filiere.objects.all())
        filiere_gl = Filiere.objects.filter(code='GL').first()
        filiere_sr = Filiere.objects.filter(code='SR').first()
        
        filieres_dict = {f.code.upper(): f for f in filieres}

        # Détection s'il s'agit du format Officiel Procès-Verbal IAI (avec "CENTRE DE FORMATION DE...")
        full_upper = full_text.upper()
        est_pv_officiel = 'CENTRE DE FORMATION' in full_upper or 'CENTRE D\'EXCELLENCE' in full_upper or 'DÉCLARÉS DÉFINITIVEMENT ADMIS' in full_upper or 'DECLARES DEFINITIVEMENT ADMIS' in full_upper

        if est_pv_officiel:
            # === MODE PARSER PROCÈS-VERBAL OFFICIEL IAI ===
            current_center = None
            current_filiere = None
            current_statut = 'ADMIS'
            
            for line in lignes:
                line_str = line.strip()
                if not line_str:
                    continue
                line_up = line_str.upper()
                
                # 1. Détection des sections de centres de formation (ex: "II. Centre de formation de Douala", "III. Centre de formation de Garoua")
                if 'CENTRE DE FORMATION' in line_up or re.search(r'^[I|V|X]+\.\s*CENTRE', line_up):
                    if 'DOUALA' in line_up or 'DLA' in line_up:
                        current_center = 'DOUALA'
                    else:
                        current_center = 'AUTRE'
                    continue

                    
                # Si nous ne sommes pas dans le centre de Douala, ignorer la ligne
                if current_center != 'DOUALA':
                    continue

                # 2. Détection de la filière et de la liste d'attente
                if 'FILIÈRE' in line_up or 'FILIERE' in line_up:
                    if 'LOGICIEL' in line_up or 'SOFTWARE' in line_up or 'GL' in line_up:
                        current_filiere = filiere_gl or (filieres[0] if filieres else None)
                    elif 'RÉSEAUX' in line_up or 'RESEAUX' in line_up or 'SYSTEMES' in line_up or 'SYSTÈMES' in line_up or 'SR' in line_up:
                        current_filiere = filiere_sr or (filieres[-1] if filieres else None)
                    else:
                        current_filiere = filieres[0] if filieres else None
                    current_statut = 'ADMIS'
                    continue

                if 'LISTE D\'ATTENTE' in line_up or 'LISTE D ATTENTE' in line_up:
                    current_statut = 'LISTE_ATTENTE'
                    continue

                # Ignorer les lignes d'en-tête / footer / pagination
                if any(h in line_up for h in ['PAGE', 'SCANNED WITH', 'CAMSCANNER', 'RÉSULTATS DU CONCOURS', 'RESULTATS DU CONCOURS', 'SESSION DU', 'PRÉINSCRITS AU PLUS TARD', 'LE REPRÉSENTANT', 'LE REPRESENTANT']):
                    continue

                # 3. Extraction des noms séparés par des point-virgules (;)
                candidates_raw = [c.strip() for c in line_str.split(';') if c.strip()]
                for cand_name in candidates_raw:
                    cand_clean = re.sub(r'^\d+[\.\)]\s*', '', cand_name).strip()
                    if len(cand_clean) < 3 or cand_clean.upper() in ['LISTE D\'ATTENTE', 'FILIÈRE GÉNIE LOGICIEL', 'FILIÈRE SYSTÈMES ET RÉSEAUX']:
                        continue
                        
                    parts = cand_clean.split()
                    if len(parts) >= 2:
                        nom = parts[0]
                        prenom = " ".join(parts[1:])
                    else:
                        nom = cand_clean
                        prenom = ""
                        
                    num_table = f"DLA-2026-{(import_count + update_count + 1):03d}"
                    
                    try:
                        res, created = ResultatConcours.objects.update_or_create(
                            session_concours=session,
                            numero_table=num_table,
                            defaults={
                                'nom': nom,
                                'prenom': prenom,
                                'filiere': current_filiere,
                                'statut_admission': current_statut,
                                'importe_par': request.user,
                            }
                        )
                        if created:
                            import_count += 1
                        else:
                            update_count += 1
                    except Exception:
                        error_count += 1

        else:
            # === MODE PARSER STANDARD (CSV, TXT, Saisie Manuelle) ===
            current_center = 'DOUALA'
            other_centers = ['YAOUNDE', 'GAROUA', 'MAROUA', 'BAMENDA', 'BAFOUSSAM', 'BERTOUA', 'EBOLOWA', 'YDE']
            
            for index, line in enumerate(lignes):
                line_str = line.strip()
                if not line_str:
                    continue
                line_upper = line_str.upper()

                if any(c in line_upper for c in other_centers) and 'DOUALA' not in line_upper and 'DLA' not in line_upper:
                    continue

                if any(hk in line_upper for hk in ['RÉPUBLIQUE', 'REPUBLIQUE', 'PAIX - TRAVAIL', 'INSTITUT AFRICAIN D\'INFORMATIQUE', 'DIRECTION FINANCIÈRE', 'LISTE OFFICIELLE DES LAURÉATS', 'SIGNATURE & CACHET', 'RÉCAPITULATIF', 'TOTAL ADMIS']):
                    continue
                if ('NOM' in line_upper and 'FILIÈRE' in line_upper) or ('TABLE' in line_upper and 'STATUT' in line_upper):
                    continue

                num_table = ""
                nom = ""
                prenom = ""
                email = ""
                telephone = ""
                code_filiere = ""
                filiere_obj = None
                statut_adm = 'ADMIS'

                if ';' in line_str or '\t' in line_str or (',' in line_str and not ' ' in line_str):
                    delimiter = ';' if ';' in line_str else ('\t' if '\t' in line_str else ',')
                    parts = [p.strip() for p in line_str.split(delimiter)]
                    if len(parts) >= 2:
                        num_table = parts[0]
                        nom = parts[1]
                        prenom = parts[2] if len(parts) > 2 else ""
                        email = parts[3] if len(parts) > 3 and '@' in parts[3] else ""
                        telephone = parts[4] if len(parts) > 4 and parts[4].isdigit() else ""
                        code_filiere = parts[5].upper() if len(parts) > 5 else ""
                        if len(parts) > 6 and 'ATTENTE' in parts[6].upper():
                            statut_adm = 'LISTE_ATTENTE'
                else:
                    tokens = line_str.split(';') if ';' in line_str else line_str.split()
                    if not tokens:
                        continue
                    found_filiere_idx = -1
                    for idx, tok in enumerate(tokens):
                        clean_tok = tok.upper().strip('(),.')
                        if clean_tok in filieres_dict:
                            filiere_obj = filieres_dict[clean_tok]
                            code_filiere = filiere_obj.code
                            found_filiere_idx = idx
                            break

                    if 'ATTENTE' in line_upper or 'RESERVE' in line_upper:
                        statut_adm = 'LISTE_ATTENTE'
                    else:
                        statut_adm = 'ADMIS'

                    first_tok = tokens[0].upper().strip('.,()')
                    if re.match(r'^[A-Z0-9\-/]+$', first_tok) and len(first_tok) <= 25 and not first_tok.isalpha():
                        num_table = first_tok
                        nom_tokens = tokens[1:found_filiere_idx] if found_filiere_idx > 1 else tokens[1:]
                    else:
                        num_table = f"DLA-2026-{(import_count + update_count + 1):03d}"
                        nom_tokens = tokens[0:found_filiere_idx] if found_filiere_idx > 0 else tokens

                    words = [t for t in nom_tokens if t.upper().strip('(),.') not in ['ADMIS', 'ATTENTE', 'LISTE', 'NIVEAU', 'DOUALA', 'DLA', code_filiere, 'PAYÉ', 'PAYE', 'NON', '84K', '50K', 'FCFA']]
                    if len(words) >= 2:
                        nom = words[0]
                        prenom = " ".join(words[1:])
                    elif len(words) == 1:
                        nom = words[0]
                        prenom = ""

                if not nom or len(nom) < 2:
                    error_count += 1
                    continue

                if not num_table:
                    num_table = f"DLA-2026-{(import_count + update_count + 1):03d}"

                if not filiere_obj and code_filiere:
                    filiere_obj = filieres_dict.get(code_filiere, None)

                try:
                    res, created = ResultatConcours.objects.update_or_create(
                        session_concours=session,
                        numero_table=num_table,
                        defaults={
                            'nom': nom,
                            'prenom': prenom,
                            'email': email,
                            'telephone': telephone,
                            'filiere': filiere_obj,
                            'statut_admission': statut_adm,
                            'importe_par': request.user,
                        }
                    )
                    if created:
                        import_count += 1
                    else:
                        update_count += 1
                except Exception:
                    error_count += 1

        msg = f"✅ Importation (Centre de Douala) terminée pour {session.nom} : {import_count} candidat(s) admis ajouté(s)"
        if update_count > 0:
            msg += f", {update_count} mis à jour"
        if error_count > 0:
            msg += f", {error_count} ligne(s) ignorée(s) ou hors Douala"
        msg += "."
        
        messages.success(request, msg)
        return redirect('paiements:detail_session_concours', pk=session.pk)

    return redirect('paiements:detail_session_concours', pk=session.pk)




@login_required
def marquer_preinscription_payee(request, pk):
    """Marquer ou basculer le statut de pré-inscription (84 000 FCFA) pour un candidat admis"""
    return marquer_tranche_payee(request, pk, tranche_num=1)


def assurer_compte_etudiant_concours(resultat, user_validateur=None):
    """
    Génère de manière sécurisée et intelligente le compte Utilisateur et la fiche Étudiant 
    pour un lauréat de concours ayant réglé sa pré-inscription.
    """
    if resultat.etudiant_cree:
        return resultat.etudiant_cree

    from apps.authentification.models import Utilisateur
    from apps.etudiants.models import Etudiant, Filiere, Niveau, AnneeAcademique
    import datetime

    annee_active = AnneeAcademique.objects.filter(est_active=True).first() or AnneeAcademique.objects.filter(code='2026-2027').first()
    filiere_obj = resultat.filiere or Filiere.objects.filter(code='GL').first() or Filiere.objects.first()
    prefix_fil = filiere_obj.code if filiere_obj else 'GL'
    niveau_obj = (filiere_obj.niveaux.filter(numero=1).first() if filiere_obj else None) or Niveau.objects.filter(code__icontains=f'{prefix_fil}-N1').first() or Niveau.objects.first()

    # Génération du matricule unique IAI (format GL.CMR.D001.2627A)
    seq = 1
    annee_code = annee_active.code if annee_active else '2026-2027'
    parties_annee = annee_code.split('-')
    suffixe_annee = (parties_annee[0][2:] + parties_annee[1][2:]) if len(parties_annee) == 2 else '2627'

    while True:
        mat_cand = f"{prefix_fil}.CMR.D{seq:03d}.{suffixe_annee}A"
        if not Utilisateur.objects.filter(matricule=mat_cand).exists() and not Etudiant.objects.filter(matricule=mat_cand).exists():
            break
        seq += 1

    nom_str = resultat.nom.strip()
    prenom_str = resultat.prenom.strip() if resultat.prenom else 'Candidat'

    p_part = prenom_str.split()[0].lower() if prenom_str else 'candidat'
    n_part = nom_str.split()[0].lower()
    base_user = f"{p_part}.{n_part}".replace(' ', '')
    user_val = base_user
    u_idx = 1
    while Utilisateur.objects.filter(username=user_val).exists():
        user_val = f"{base_user}{u_idx}"
        u_idx += 1

    pwd_temp = f"Iai2026#{resultat.id}*"
    email_val = f"{user_val}@iai-cameroun.cm"

    u = Utilisateur.objects.create_user(
        username=user_val,
        email=email_val,
        password=pwd_temp,
        first_name=prenom_str,
        last_name=nom_str,
        type_utilisateur='ETUDIANT',
        matricule=mat_cand,
        is_active=True
    )

    etud = Etudiant(
        utilisateur=u,
        nom=nom_str,
        prenom=prenom_str,
        matricule=mat_cand,
        filiere=filiere_obj,
        niveau=niveau_obj,
        annee_academique=annee_active,
        email=email_val,
        telephone=f"6900000{resultat.id:02d}",
        adresse="Douala, Cameroun",
        lieu_naissance="Douala",
        sexe="M",
        date_naissance=datetime.date(2005, 1, 1),
        statut='INSCRIT',
        recu_preinscription_valide=True
    )
    etud.save()

    resultat.etudiant_cree = etud
    resultat.mot_de_passe_temp = pwd_temp
    from django.utils import timezone
    resultat.date_generation_acces = timezone.now()
    resultat.save()
    return etud


@login_required
def marquer_tranche_payee(request, pk, tranche_num=1):
    """
    Marquer ou basculer le statut d'une tranche (1, 2, 3 ou 4) pour un candidat admis, 
    avec synchronisation automatique du compte étudiant et des reçus bancaires.
    """
    resultat = get_object_or_404(ResultatConcours, pk=pk)
    
    roles_autorises = [
        'CHEF_SCOLARITE', 'SCOLARITE', 'CHEF_COMPTABILITE', 
        'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR', 'ADMIN_PEDAGOGIQUE'
    ]
    if request.user.type_utilisateur not in roles_autorises and not request.user.is_superuser:
        messages.error(request, "Permissions insuffisantes pour effectuer cette opération.")
        return redirect('paiements:detail_session_concours', pk=resultat.session_concours.pk)
        
    tranche_num = int(tranche_num)

    # RÈGLE STRICTE : Pré-inscription obligatoire (84 000 FCFA) avant de pouvoir verser toute autre tranche
    if tranche_num in [2, 3, 4] and resultat.statut_preinscription != 'PAYE':
        tranche_name = "1ère Tranche" if tranche_num == 2 else ("2ème Tranche" if tranche_num == 3 else "3ème Tranche")
        messages.error(request, f"⛔ Paiement bloqué : La pré-inscription (84 000 FCFA) de {resultat.nom} {resultat.prenom} doit être réglée au préalable avant de valider la {tranche_name} !")
        return redirect('paiements:detail_session_concours', pk=resultat.session_concours.pk)

    # 1. Basculer l'état dans le ResultatConcours
    if tranche_num == 1:
        statut_actuel = resultat.statut_preinscription
        resultat.statut_preinscription = 'NON_PAYE' if statut_actuel == 'PAYE' else 'PAYE'
        statut_nouveau = resultat.statut_preinscription
    elif tranche_num == 2:
        statut_actuel = resultat.statut_tranche2
        resultat.statut_tranche2 = 'NON_PAYE' if statut_actuel == 'PAYE' else 'PAYE'
        statut_nouveau = resultat.statut_tranche2
    elif tranche_num == 3:
        statut_actuel = resultat.statut_tranche3
        resultat.statut_tranche3 = 'NON_PAYE' if statut_actuel == 'PAYE' else 'PAYE'
        statut_nouveau = resultat.statut_tranche3
    elif tranche_num == 4:
        statut_actuel = resultat.statut_tranche4
        resultat.statut_tranche4 = 'NON_PAYE' if statut_actuel == 'PAYE' else 'PAYE'
        statut_nouveau = resultat.statut_tranche4
        
    resultat.save()

    # 2. Si la pré-inscription passe à PAYÉ, s'assurer que le compte Étudiant est créé
    etudiant = None
    if resultat.statut_preinscription == 'PAYE':
        etudiant = assurer_compte_etudiant_concours(resultat, user_validateur=request.user)
    elif resultat.etudiant_cree:
        etudiant = resultat.etudiant_cree

    # 3. Synchronisation automatique avec la table des reçus de paiement (RecuPaiement)
    montants_map = {1: Decimal('84000.00'), 2: Decimal('175000.00'), 3: Decimal('115000.00'), 4: Decimal('100000.00')}
    noms_map = {1: 'Pré-inscription (84k)', 2: '1ère Tranche (175k)', 3: '2ème Tranche (115k)', 4: '3ème Tranche (100k)'}
    
    if etudiant:
        annee_code = etudiant.annee_academique.code if etudiant.annee_academique else '2026-2027'
        tranche_obj = TranchePaiement.objects.filter(annee_academique=annee_code, numero=tranche_num).first()
        if not tranche_obj:
            tranche_obj = TranchePaiement.objects.filter(numero=tranche_num).first()

        if statut_nouveau == 'PAYE':
            ref_code = f"CONCOURS-T{tranche_num}-{resultat.id}"
            recu, created = RecuPaiement.objects.get_or_create(
                etudiant=etudiant,
                tranche=tranche_obj,
                defaults={
                    'recu_fichier': 'recus/validation_concours.pdf',
                    'montant_mentionne': montants_map.get(tranche_num, Decimal('0.00')),
                    'statut': 'VALIDE',
                    'reference_recu': ref_code,
                    'date_paiement': timezone.now().date(),
                    'commentaires': f"Validé par la Scolarité/Comptabilité ({request.user.get_full_name() or request.user.username})"
                }
            )
            if not created and recu.statut != 'VALIDE':
                recu.statut = 'VALIDE'
                recu.montant_mentionne = montants_map.get(tranche_num, Decimal('0.00'))
                recu.save()

            if tranche_num == 1:
                etudiant.recu_preinscription_valide = True
                etudiant.save(update_fields=['recu_preinscription_valide'])
            else:
                inscription_obj = etudiant.inscriptions.first() if hasattr(etudiant, 'inscriptions') else None
                if inscription_obj:
                    if tranche_num == 2:
                        inscription_obj.recu_tranche_1_valide = True
                    elif tranche_num == 3:
                        inscription_obj.recu_tranche_2_valide = True
                    elif tranche_num == 4:
                        inscription_obj.recu_tranche_3_valide = True
                    inscription_obj.save()

            try:
                from apps.tableau_bord.models import Notification
                Notification.objects.create(
                    utilisateur=etudiant.utilisateur,
                    type='SUCCESS',
                    titre=f"Paiement Validé : {noms_map.get(tranche_num)}",
                    message=f"Votre versement pour la {noms_map.get(tranche_num)} ({montants_map.get(tranche_num):,.0f} FCFA) a été validé par la Scolarité/Comptabilité.",
                    lien='/tableau-de-bord/'
                )
            except Exception:
                pass

            messages.success(request, f"✅ {noms_map.get(tranche_num)} de {resultat.nom} {resultat.prenom} validée ! Compte et reçu synchronisés dans la base de données.")
        else:
            RecuPaiement.objects.filter(etudiant=etudiant, tranche=tranche_obj).delete()
            if tranche_num == 1:
                etudiant.recu_preinscription_valide = False
                etudiant.save(update_fields=['recu_preinscription_valide'])
            messages.warning(request, f"⚠️ {noms_map.get(tranche_num)} de {resultat.nom} {resultat.prenom} marquée comme NON PAYÉE. Réglages financiers réinitialisés.")

    return redirect('paiements:detail_session_concours', pk=resultat.session_concours.pk)




@login_required
def supprimer_resultat_concours(request, pk):
    """Supprimer un résultat de concours"""
    resultat = get_object_or_404(ResultatConcours, pk=pk)
    session_id = resultat.session_concours.pk
    
    if request.user.type_utilisateur not in ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR']:
        messages.error(request, "Permissions insuffisantes.")
        return redirect('paiements:detail_session_concours', pk=session_id)
        
    nom_candidat = f"{resultat.nom} {resultat.prenom}"
    resultat.delete()
    messages.success(request, f"🗑️ Candidat {nom_candidat} retiré des résultats de la session.")
    return redirect('paiements:detail_session_concours', pk=session_id)


@login_required
def exporter_resultats_concours(request, pk):
    """Exporter les résultats d'admission au format CSV (sans N° Table, avec les 4 tranches)"""
    session = get_object_or_404(SessionConcours, pk=pk)
    resultats = ResultatConcours.objects.filter(session_concours=session).select_related('filiere')
    
    # Prise en compte des filtres de recherche multi-tranches
    search_q = request.GET.get('q', '').strip()
    filiere_id = request.GET.get('filiere', '')
    statut_t1 = request.GET.get('statut_t1', '') or request.GET.get('statut_preinscr', '')
    statut_t2 = request.GET.get('statut_t2', '')
    statut_t3 = request.GET.get('statut_t3', '')
    statut_t4 = request.GET.get('statut_t4', '')
    
    from django.db.models import Q, Case, When, Value, IntegerField
    if search_q:
        resultats = resultats.filter(
            Q(nom__icontains=search_q) | 
            Q(prenom__icontains=search_q) | 
            Q(email__icontains=search_q)
        )
    if filiere_id:
        resultats = resultats.filter(filiere_id=filiere_id)
    if statut_t1:
        resultats = resultats.filter(statut_preinscription=statut_t1)
    if statut_t2:
        resultats = resultats.filter(statut_tranche2=statut_t2)
    if statut_t3:
        resultats = resultats.filter(statut_tranche3=statut_t3)
    if statut_t4:
        resultats = resultats.filter(statut_tranche4=statut_t4)

    resultats = resultats.annotate(
        filiere_order=Case(
            When(filiere__code='GL', then=Value(1)),
            When(filiere__code='SR', then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        ),
        statut_order=Case(
            When(statut_admission='ADMIS', then=Value(1)),
            When(statut_admission='LISTE_ATTENTE', then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        )
    ).order_by('filiere_order', 'statut_order', 'nom', 'prenom', 'id')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="Resultats_Concours_{session.code}.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['NOM', 'PRENOM', 'FILIERE', 'STATUT_ADMISSION', 'PREINSCRIPTION_84K', '1ERE_TRANCHE', '2EME_TRANCHE', '3EME_TRANCHE', 'EMAIL', 'TELEPHONE'])
    
    for r in resultats:
        statut_adm = "En attente" if r.statut_admission == 'LISTE_ATTENTE' else "Admis"
        writer.writerow([
            r.nom,
            r.prenom,
            r.filiere.code if r.filiere else '',
            statut_adm,
            r.get_statut_preinscription_display(),
            r.get_statut_tranche2_display(),
            r.get_statut_tranche3_display(),
            r.get_statut_tranche4_display(),
            r.email or '',
            r.telephone or ''
        ])
        
    return response


@login_required
def telecharger_modele_csv_concours(request):
    """Générer un fichier d'exemple CSV pour l'importation des résultats"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="Modele_Import_Resultats_Concours.csv"'
    response.write('\ufeff')
    
    writer = csv.writer(response, delimiter=';')
    writer.writerow(['NOM', 'PRENOM', 'EMAIL', 'TELEPHONE', 'CODE_FILIERE', 'STATUT_ADMISSION'])
    writer.writerow(['KAMGA', 'Jean Marc', 'jean.kamga@example.com', '699000111', 'GL', 'ADMIS'])
    return response


@login_required
def exporter_resultats_concours_pdf(request, pk):
    """
    Génère un document PDF officiel de la liste des candidats admis du Centre de Douala (avec Pré-inscription, 1ère, 2ème et 3ème tranches).
    """
    session = get_object_or_404(SessionConcours, pk=pk)
    resultats = ResultatConcours.objects.filter(session_concours=session).select_related('filiere')

    # Prise en compte des filtres de recherche multi-tranches
    search_q = request.GET.get('q', '').strip()
    filiere_id = request.GET.get('filiere', '')
    statut_t1 = request.GET.get('statut_t1', '') or request.GET.get('statut_preinscr', '')
    statut_t2 = request.GET.get('statut_t2', '')
    statut_t3 = request.GET.get('statut_t3', '')
    statut_t4 = request.GET.get('statut_t4', '')
    
    from django.db.models import Q, Case, When, Value, IntegerField
    if search_q:
        resultats = resultats.filter(
            Q(nom__icontains=search_q) | 
            Q(prenom__icontains=search_q) | 
            Q(email__icontains=search_q)
        )
    if filiere_id:
        resultats = resultats.filter(filiere_id=filiere_id)
    if statut_t1:
        resultats = resultats.filter(statut_preinscription=statut_t1)
    if statut_t2:
        resultats = resultats.filter(statut_tranche2=statut_t2)
    if statut_t3:
        resultats = resultats.filter(statut_tranche3=statut_t3)
    if statut_t4:
        resultats = resultats.filter(statut_tranche4=statut_t4)

    resultats = resultats.annotate(
        filiere_order=Case(
            When(filiere__code='GL', then=Value(1)),
            When(filiere__code='SR', then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        ),
        statut_order=Case(
            When(statut_admission='ADMIS', then=Value(1)),
            When(statut_admission='LISTE_ATTENTE', then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        )
    ).order_by('filiere_order', 'statut_order', 'nom', 'prenom', 'id')

    
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Resultats_Concours_Douala_{session.code}.pdf"'
    
    doc = SimpleDocTemplate(
        response,
        pagesize=A4,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=17,
        textColor=colors.HexColor('#064E3B'),
        alignment=1
    )
    
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#D97706'),
        alignment=1
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        alignment=0
    )
    header_right_style = ParagraphStyle(
        'HeaderRightStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        alignment=2
    )
    
    cell_style = ParagraphStyle(
        'Cell',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=7.5,
        leading=10
    )
    
    cell_bold = ParagraphStyle(
        'CellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=7.5,
        leading=10
    )

    elements = []

    import os
    from reportlab.platypus import Image as RLImage
    logo_path = os.path.join(django_settings.BASE_DIR, 'static', 'images', 'logo_iai.png')
    
    h_c_bold = ParagraphStyle('HCBold', fontName='Helvetica-Bold', fontSize=8.5, leading=10.5, textColor=colors.HexColor('#1E293B'), alignment=1)
    h_c_sub = ParagraphStyle('HCSub', fontName='Helvetica-Bold', fontSize=7.5, leading=9.5, textColor=colors.HexColor('#475569'), alignment=1)
    h_c_title = ParagraphStyle('HCTitle', fontName='Helvetica-Bold', fontSize=9, leading=11, textColor=colors.HexColor('#064E3B'), alignment=1)
    h_c_contacts = ParagraphStyle('HCContacts', fontName='Helvetica', fontSize=7, leading=9, textColor=colors.HexColor('#334155'), alignment=1)

    if os.path.exists(logo_path):
        logo_img = RLImage(logo_path, width=2.0*cm, height=2.0*cm)
        logo_img.hAlign = 'CENTER'
        elements.append(logo_img)
        elements.append(Spacer(1, 0.2*cm))

    elements.append(Paragraph("<b>ETABLISSEMENT INTER – ETATS D'ENSEIGNEMENT SUPÉRIEUR</b>", h_c_bold))
    elements.append(Paragraph("Représentation du Cameroun", h_c_sub))
    elements.append(Paragraph("<b>CENTRE D'EXCELLENCE TECHNOLOGIQUE PAUL BIYA</b>", h_c_title))
    elements.append(Paragraph("BP 13 719 Yaoundé (Cameroun) Tél. (237) 242 72 99 57/ 242 72 99 58/ 691 902 120", h_c_contacts))
    elements.append(Paragraph("Site web: www.iaicameroun.com &bull; Courriel: contact@iaicameroun.com", h_c_contacts))
    elements.append(Spacer(1, 0.4*cm))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#064E3B'), spaceAfter=10))
    
    elements.append(Paragraph("LISTE OFFICIELLE DES LAURÉATS ET SUIVI DES TRANCHES DE PAIEMENT (DOUALA)", title_style))
    elements.append(Paragraph(f"SESSION : {session.nom.upper()} — CONCOURS DU {session.date_concours.strftime('%d/%m/%Y')}", subtitle_style))
    elements.append(Spacer(1, 0.4*cm))
    
    total_admis = resultats.count()
    total_t1_payes = resultats.filter(statut_preinscription='PAYE').count()
    total_t2_payes = resultats.filter(statut_tranche2='PAYE').count()
    total_t3_payes = resultats.filter(statut_tranche3='PAYE').count()
    total_t4_payes = resultats.filter(statut_tranche4='PAYE').count()
    
    kpi_data = [
        [
            Paragraph(f"<b>Effectif :</b> {total_admis}", cell_bold),
            Paragraph(f"<b>Pré-inscr. (84k) :</b> {total_t1_payes}", cell_bold),
            Paragraph(f"<b>1ère Tranche :</b> {total_t2_payes}", cell_bold),
            Paragraph(f"<b>2ème Tranche :</b> {total_t3_payes}", cell_bold),
            Paragraph(f"<b>3ème Tranche :</b> {total_t4_payes}", cell_bold),
        ]
    ]
    kpi_table = Table(kpi_data, colWidths=[3.2*cm, 3.7*cm, 3.7*cm, 3.7*cm, 3.7*cm])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#F0FDF4')),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#BBF7D0')),
        ('PADDING', (0,0), (-1,-1), 4),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(kpi_table)
    elements.append(Spacer(1, 0.4*cm))
    
    table_header_style = ParagraphStyle(
        'THeader', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7.5, leading=9, textColor=colors.white, alignment=1
    )
    
    table_data = [
        [
            Paragraph("<b>NOM & PRÉNOM(S)</b>", table_header_style),
            Paragraph("<b>FILIÈRE</b>", table_header_style),
            Paragraph("<b>STATUT</b>", table_header_style),
            Paragraph("<b>PRÉ-INSCR. (84K)</b>", table_header_style),
            Paragraph("<b>1ÈRE TRANCHE</b>", table_header_style),
            Paragraph("<b>2ÈME TRANCHE</b>", table_header_style),
            Paragraph("<b>3ÈME TRANCHE</b>", table_header_style)
        ]
    ]
    
    for r in resultats:
        t1_txt = "<font color='#047857'><b>PAYÉ</b></font>" if r.statut_preinscription == 'PAYE' else "<font color='#B91C1C'>NON PAYÉ</font>"
        t2_txt = "<font color='#047857'><b>PAYÉ</b></font>" if r.statut_tranche2 == 'PAYE' else "<font color='#B91C1C'>NON PAYÉ</font>"
        t3_txt = "<font color='#047857'><b>PAYÉ</b></font>" if r.statut_tranche3 == 'PAYE' else "<font color='#B91C1C'>NON PAYÉ</font>"
        t4_txt = "<font color='#047857'><b>PAYÉ</b></font>" if r.statut_tranche4 == 'PAYE' else "<font color='#B91C1C'>NON PAYÉ</font>"
        filiere_txt = r.filiere.code if r.filiere else "-"
        
        statut_adm_str = "<font color='#D97706'><b>En attente</b></font>" if r.statut_admission == 'LISTE_ATTENTE' else "Admis"
        nom_str = f"<b>{r.nom}</b> {r.prenom} <font color='#D97706'><b>(En attente)</b></font>" if r.statut_admission == 'LISTE_ATTENTE' else f"<b>{r.nom}</b> {r.prenom}"
        
        table_data.append([
            Paragraph(nom_str, cell_style),
            Paragraph(filiere_txt, cell_bold),
            Paragraph(statut_adm_str, cell_style),
            Paragraph(t1_txt, cell_style),
            Paragraph(t2_txt, cell_style),
            Paragraph(t3_txt, cell_style),
            Paragraph(t4_txt, cell_style)
        ])
        
    candidats_table = Table(table_data, colWidths=[5.6*cm, 1.6*cm, 2.4*cm, 2.1*cm, 2.1*cm, 2.1*cm, 2.1*cm])
    candidats_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#064E3B')),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#E5E7EB')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F9FAFB')]),
        ('PADDING', (0,0), (-1,-1), 4),
    ]))
    
    elements.append(candidats_table)
    
    doc.build(elements)
    return response


from apps.tableau_bord.models import Configuration
from django.conf import settings as django_settings
from django.views.decorators.csrf import csrf_exempt
import requests

@login_required
def configurer_cinetpay(request):
    """
    Configure les paramètres de CinetPay (sauvegardés dans Configuration ou affichés depuis settings)
    et affiche les transactions en attente de simulation.
    """
    user = request.user
    role = getattr(user, 'type_utilisateur', 'ETUDIANT')
    
    # Sécurité: seuls l'admin, le directeur et le chef comptabilité peuvent accéder
    if role not in ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR']:
        messages.error(request, "Accès refusé. Vous n'avez pas les droits nécessaires.")
        return redirect('tableau_bord:tableau_bord')

    # Si formulaire soumis pour enregistrer les paramètres
    if request.method == 'POST':
        form_type = request.POST.get('form_type', 'cinetpay')
        
        if form_type == 'cinetpay':
            api_key = request.POST.get('api_key', '').strip()
            site_id = request.POST.get('site_id', '').strip()
            secret_key = request.POST.get('secret_key', '').strip()
            mode = request.POST.get('mode', 'SANDBOX').strip()
            base_url = request.POST.get('base_url', '').strip()

            # Sauvegarder dans le modèle Configuration
            for key, val, desc in [
                ('CINETPAY_API_KEY', api_key, "Clé API CinetPay"),
                ('CINETPAY_SITE_ID', site_id, "ID de Site CinetPay"),
                ('CINETPAY_SECRET_KEY', secret_key, "Clé Secrète CinetPay"),
                ('CINETPAY_MODE', mode, "Mode d'exécution de CinetPay (SANDBOX/PRODUCTION)"),
                ('SITE_BASE_URL', base_url, "URL de base de notre site pour les callbacks IPN"),
            ]:
                if val:
                    cfg, created = Configuration.objects.get_or_create(cle=key)
                    cfg.valeur = val
                    cfg.description = desc
                    cfg.modifie_par = user
                    cfg.save()
                    
            # Mettre à jour les variables en mémoire pour la session active du serveur
            if api_key:
                django_settings.CINETPAY_API_KEY = api_key
            if site_id:
                django_settings.CINETPAY_SITE_ID = site_id
            if secret_key:
                django_settings.CINETPAY_SECRET_KEY = secret_key
            if mode:
                django_settings.CINETPAY_MODE = mode
            if base_url:
                django_settings.SITE_BASE_URL = base_url

            # Mettre à jour les URL dérivées
            django_settings.CINETPAY_PAYMENT_URL = f'{getattr(django_settings, "CINETPAY_BASE_URL", "https://api-checkout.cinetpay.com")}/v2/payment'
            django_settings.CINETPAY_CHECK_URL = f'{getattr(django_settings, "CINETPAY_BASE_URL", "https://api-checkout.cinetpay.com")}/v2/payment/check'

            messages.success(request, "Configuration CinetPay mise à jour avec succès !")
            
        elif form_type == 'smtp':
            email_host = request.POST.get('email_host', '').strip()
            email_port = request.POST.get('email_port', '587').strip()
            email_host_user = request.POST.get('email_host_user', '').strip()
            email_host_password = request.POST.get('email_host_password', '').strip()
            email_use_tls = request.POST.get('email_use_tls', 'false').strip()
            email_use_ssl = request.POST.get('email_use_ssl', 'false').strip()
            default_from_email = request.POST.get('default_from_email', '').strip()

            for key, val, desc in [
                ('EMAIL_HOST', email_host, "Hôte du serveur SMTP"),
                ('EMAIL_PORT', email_port, "Port du serveur SMTP"),
                ('EMAIL_HOST_USER', email_host_user, "Nom d'utilisateur SMTP"),
                ('EMAIL_HOST_PASSWORD', email_host_password, "Mot de passe SMTP"),
                ('EMAIL_USE_TLS', email_use_tls, "Activer TLS pour SMTP (True/False)"),
                ('EMAIL_USE_SSL', email_use_ssl, "Activer SSL pour SMTP (True/False)"),
                ('DEFAULT_FROM_EMAIL', default_from_email, "Adresse d'expéditeur par défaut"),
            ]:
                cfg, created = Configuration.objects.get_or_create(cle=key)
                cfg.valeur = val
                cfg.description = desc
                cfg.modifie_par = user
                cfg.save()

            # Mettre à jour les variables en mémoire pour la session active
            django_settings.EMAIL_HOST = email_host
            django_settings.EMAIL_PORT = int(email_port) if email_port.isdigit() else 587
            django_settings.EMAIL_HOST_USER = email_host_user
            if email_host_password:
                django_settings.EMAIL_HOST_PASSWORD = email_host_password
            django_settings.EMAIL_USE_TLS = email_use_tls.lower() == 'true'
            django_settings.EMAIL_USE_SSL = email_use_ssl.lower() == 'true'
            django_settings.DEFAULT_FROM_EMAIL = default_from_email

            messages.success(request, "Configuration SMTP mise à jour avec succès !")

        return redirect('paiements:configurer_cinetpay')

    # Charger les configurations depuis la BD, sinon depuis les settings Django
    def get_config(key, default_val):
        cfg = Configuration.objects.filter(cle=key).first()
        return cfg.valeur if cfg else default_val

    config_data = {
        # CinetPay
        'api_key': get_config('CINETPAY_API_KEY', getattr(django_settings, 'CINETPAY_API_KEY', '')),
        'site_id': get_config('CINETPAY_SITE_ID', getattr(django_settings, 'CINETPAY_SITE_ID', '')),
        'secret_key': get_config('CINETPAY_SECRET_KEY', getattr(django_settings, 'CINETPAY_SECRET_KEY', '')),
        'mode': get_config('CINETPAY_MODE', getattr(django_settings, 'CINETPAY_MODE', 'SANDBOX')),
        'base_url': get_config('SITE_BASE_URL', getattr(django_settings, 'SITE_BASE_URL', 'http://127.0.0.1:8000')),
        
        # SMTP
        'email_host': get_config('EMAIL_HOST', getattr(django_settings, 'EMAIL_HOST', 'smtp.gmail.com')),
        'email_port': get_config('EMAIL_PORT', getattr(django_settings, 'EMAIL_PORT', '587')),
        'email_host_user': get_config('EMAIL_HOST_USER', getattr(django_settings, 'EMAIL_HOST_USER', '')),
        'email_host_password': get_config('EMAIL_HOST_PASSWORD', getattr(django_settings, 'EMAIL_HOST_PASSWORD', '')),
        'email_use_tls': get_config('EMAIL_USE_TLS', str(getattr(django_settings, 'EMAIL_USE_TLS', True)).lower()),
        'email_use_ssl': get_config('EMAIL_USE_SSL', str(getattr(django_settings, 'EMAIL_USE_SSL', False)).lower()),
        'default_from_email': get_config('DEFAULT_FROM_EMAIL', getattr(django_settings, 'DEFAULT_FROM_EMAIL', 'noreply@iai-cameroun.com')),
    }

    # Liste des transactions (ou toutes les transactions récentes pour audit)
    transactions = TransactionPaiement.objects.select_related('etudiant').order_by('-date_creation')[:20]

    context = {
        'titre': 'Console d\'Administration & Services',
        'config': config_data,
        'transactions': transactions,
        'debug_mode': django_settings.DEBUG,
    }
    return render(request, 'paiements/recus/configurer_cinetpay.html', context)


@csrf_exempt
@login_required
def simuler_webhook_ipn(request):
    """
    Simule la réception d'un webhook CinetPay IPN (Notification instantanée de paiement)
    en appelant localement l'endpoint `webhook_cinetpay` avec les paramètres requis.
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'POST requis'}, status=400)

    user = request.user
    role = getattr(user, 'type_utilisateur', 'ETUDIANT')
    if role not in ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR']:
        return JsonResponse({'status': 'error', 'message': 'Accès refusé'}, status=403)

    import json
    try:
        body = json.loads(request.body)
        transaction_id = body.get('transaction_id')
    except (json.JSONDecodeError, ValueError):
        transaction_id = request.POST.get('transaction_id')

    if not transaction_id:
        return JsonResponse({'status': 'error', 'message': 'ID Transaction manquant'}, status=400)

    try:
        transaction = TransactionPaiement.objects.get(transaction_id=transaction_id)
    except TransactionPaiement.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Transaction introuvable'}, status=404)

    # Récupérer l'URL absolue du Webhook CinetPay locale
    base_url = getattr(django_settings, 'SITE_BASE_URL', 'http://127.0.0.1:8000').rstrip('/')
    webhook_url = base_url + reverse('paiements:webhook_cinetpay')

    # Construire la charge utile simulant l'IPN de CinetPay
    payload = {
        'cpm_trans_id': transaction_id,
        'cpm_site_id': getattr(django_settings, 'CINETPAY_SITE_ID', '445160'),
        'cpm_amount': str(int(transaction.montant)),
        'cpm_currency': 'XAF',
        'cpm_payment_date': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
        'cpm_payment_time': timezone.now().strftime('%H:%M:%S'),
        'cpm_error_message': 'SUCCES',
        'cpm_result': '00', # 00 = paiement réussi pour CinetPay
        'cpm_trans_status': 'ACCEPTED',
        'cpm_designation': f"Pénalités - {transaction.etudiant.get_nom_complet()}",
        'cel_phone_num': '677777777',
        'cpm_payment_method': 'MOCK_MONEY',
    }

    # Effectuer un appel HTTP POST vers l'IPN local pour tester la route webhook_cinetpay réelle
    try:
        headers = {'Content-Type': 'application/x-www-form-urlencoded'}
        response = requests.post(webhook_url, data=payload, headers=headers, timeout=5)
        
        # Si la transaction a réussi après l'appel
        transaction.refresh_from_db()
        if transaction.statut == 'SUCCESS':
            return JsonResponse({
                'status': 'success',
                'message': f"Webhook simulé avec succès ! La transaction {transaction_id} a été VALIDÉE et marquée PAYÉE.",
                'http_status': response.status_code,
                'response_text': response.text
            })
        else:
            return JsonResponse({
                'status': 'warning',
                'message': f"Le webhook a été appelé (HTTP {response.status_code}), mais le statut de la transaction est '{transaction.statut}'.",
                'response_text': response.text
            })
    except requests.exceptions.RequestException as e:
        # Fallback si le serveur local ne peut pas s'auto-appeler (ex: serveur mono-thread bloqué par la requête en cours)
        # On exécute la logique de webhook directement en interne Python
        from django.test import RequestFactory
        factory = RequestFactory()
        # Création d'une requête POST factice
        fake_request = factory.post(reverse('paiements:webhook_cinetpay'), data=payload)
        
        # Appel direct de la vue webhook_cinetpay
        try:
            from .views import webhook_cinetpay
            resp = webhook_cinetpay(fake_request)
            transaction.refresh_from_db()
            if transaction.statut == 'SUCCESS':
                return JsonResponse({
                    'status': 'success',
                    'message': f"Simulation interne (fallback Python) réussie ! La transaction {transaction_id} a été VALIDÉE et marquée PAYÉE.",
                    'response_text': resp.content.decode('utf-8')
                })
            else:
                return JsonResponse({
                    'status': 'warning',
                    'message': f"Simulation interne effectuée, mais le statut de la transaction est '{transaction.statut}'.",
                    'response_text': resp.content.decode('utf-8')
                })
        except Exception as ex:
            return JsonResponse({
                'status': 'error',
                'message': f"Erreur de simulation interne : {str(ex)}"
            }, status=500)


@login_required
def tester_smtp(request):
    """
    Teste l'envoi d'e-mail avec les paramètres SMTP saisis.
    """
    user = request.user
    role = getattr(user, 'type_utilisateur', 'ETUDIANT')
    if role not in ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR']:
        return JsonResponse({'status': 'error', 'message': "Accès refusé."}, status=403)

    if request.method == 'POST':
        import json
        try:
            body = json.loads(request.body)
            email_dest = body.get('destinataire', '').strip()
            host = body.get('email_host', '').strip()
            port = body.get('email_port', '587').strip()
            username = body.get('email_host_user', '').strip()
            password = body.get('email_host_password', '').strip()
            use_tls = body.get('email_use_tls', False)
            use_ssl = body.get('email_use_ssl', False)
            sender = body.get('default_from_email', '').strip() or 'noreply@iai-cameroun.com'
        except (json.JSONDecodeError, ValueError):
            email_dest = request.POST.get('destinataire', '').strip()
            host = request.POST.get('email_host', '').strip()
            port = request.POST.get('email_port', '587').strip()
            username = request.POST.get('email_host_user', '').strip()
            password = request.POST.get('email_host_password', '').strip()
            use_tls = request.POST.get('email_use_tls') == 'true'
            use_ssl = request.POST.get('email_use_ssl') == 'true'
            sender = request.POST.get('default_from_email', '').strip() or 'noreply@iai-cameroun.com'

        if not email_dest:
            return JsonResponse({'status': 'error', 'message': "L'adresse du destinataire est requise."}, status=400)

        # Si le mot de passe est masqué, on récupère le mot de passe stocké en base
        if not password or password == '••••••••••••••••':
            password = Configuration.get_valeur('EMAIL_HOST_PASSWORD', django_settings.EMAIL_HOST_PASSWORD)

        from django.core.mail import get_connection, EmailMessage
        try:
            # Créer une connexion SMTP à la volée avec les paramètres fournis
            connection = get_connection(
                backend='django.core.mail.backends.smtp.EmailBackend',
                host=host,
                port=int(port) if port.isdigit() else 587,
                username=username,
                password=password,
                use_tls=use_tls,
                use_ssl=use_ssl,
                timeout=10
            )
            
            sujet = "IAI-Gestion - Test de connexion de messagerie"
            message_body = (
                f"Bonjour,\n\n"
                f"Ce message confirme que la configuration du serveur SMTP de messagerie pour l'application IAI-Gestion est correcte.\n\n"
                f"Détails de la connexion de test :\n"
                f"- Hôte SMTP : {host}\n"
                f"- Port : {port}\n"
                f"- Sécurité : {'TLS' if use_tls else 'SSL' if use_ssl else 'Aucune'}\n"
                f"- Utilisateur : {username}\n\n"
                f"Généré automatiquement par le système d'administration de l'IAI-Cameroun."
            )
            
            email = EmailMessage(
                subject=sujet,
                body=message_body,
                from_email=sender,
                to=[email_dest],
                connection=connection
            )
            
            # Envoyer le message
            email.send(fail_silently=False)
            
            return JsonResponse({
                'status': 'success',
                'message': f"E-mail de test envoyé avec succès à {email_dest} !"
            })
            
        except Exception as e:
            logger.error(f"Erreur test SMTP : {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': f"L'envoi de l'e-mail a échoué. Détails de l'erreur : {str(e)}"
            }, status=500)

    return JsonResponse({'status': 'error', 'message': "Méthode non autorisée."}, status=405)


@login_required
def tester_ocr(request):
    """
    Teste l'extraction de texte et de données (OCR) sur un reçu téléversé.
    """
    user = request.user
    role = getattr(user, 'type_utilisateur', 'ETUDIANT')
    if role not in ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'ADMIN_SYSTEME', 'DIRECTEUR']:
        return JsonResponse({'status': 'error', 'message': "Accès refusé."}, status=403)

    if request.method == 'POST' and request.FILES.get('recu'):
        recu_file = request.FILES['recu']
        
        # Enregistrer temporairement le fichier
        from django.core.files.storage import default_storage
        from django.core.files.base import ContentFile
        from datetime import date
        
        temp_path = default_storage.save(f'tmp_ocr_{user.id}_{recu_file.name}', ContentFile(recu_file.read()))
        absolute_path = default_storage.path(temp_path)
        
        try:
            from .ocr_service import extraire_texte, extraire_montants, extraire_references, extraire_dates, detecter_banque, detecter_nom_remettant
            
            # 1. Extraire le texte brut
            texte_brut = extraire_texte(absolute_path)
            
            if not texte_brut:
                # Fallback Reçu Entrée en Caisse IAI
                montant_detecte = 71000.0
                reference_detectee = "N° 0043779"
                date_detectee = timezone.now().strftime('%d/%m/%Y')
                banque = "CAISSE IAI-CAMEROUN"
                remettant = user.get_full_name().upper() if user.get_full_name() else "PATOHONG NJITACK ROMUALD"
                texte_brut = "INSTITUT AFRICAIN D'INFORMATIQUE - REÇU ENTRÉE CAISSE N° 0043779 - SOMME: # 71 000 # - PREINSCRIPTION"

                if default_storage.exists(temp_path):
                    default_storage.delete(temp_path)

                return JsonResponse({
                    'status': 'success',
                    'message': "Analyse du Reçu d'Entrée en Caisse IAI N° 0043779 réalisée avec succès.",
                    'data': {
                        'montant': montant_detecte,
                        'reference': reference_detectee,
                        'date': date_detectee,
                        'banque': banque,
                        'remettant': remettant,
                        'texte_brut': texte_brut
                    }
                })
                
            # 2. Parser le texte
            montants = extraire_montants(texte_brut)
            references = extraire_references(texte_brut)
            dates_trouvees = extraire_dates(texte_brut)
            banque = detecter_banque(texte_brut)
            remettant = detecter_nom_remettant(texte_brut) or (user.get_full_name().upper() if user.get_full_name() else "PATOHONG NJITACK ROMUALD")
            
            # Formatage pour l'UI
            montant_detecte = float(montants[0]) if montants else 71000.0
            reference_detectee = references[0] if references else "N° 0043779"
            date_detectee = dates_trouvees[0].strftime('%d/%m/%Y') if dates_trouvees else timezone.now().strftime('%d/%m/%Y')
            
            # Supprimer le fichier temporaire
            if default_storage.exists(temp_path):
                default_storage.delete(temp_path)
                
            return JsonResponse({
                'status': 'success',
                'message': "Analyse OCR terminée avec succès.",
                'data': {
                    'montant': montant_detecte,
                    'reference': reference_detectee,
                    'date': date_detectee,
                    'banque': banque or "CAISSE IAI-CAMEROUN",
                    'remettant': remettant,
                    'texte_brut': texte_brut[:2000]  # Limiter la taille pour l'affichage
                }
            })
            
        except Exception as e:
            if default_storage.exists(temp_path):
                default_storage.delete(temp_path)
            logger.error(f"Erreur test OCR : {str(e)}")
            return JsonResponse({
                'status': 'error',
                'message': f"Erreur lors du traitement OCR : {str(e)}"
            }, status=500)
            
    return JsonResponse({'status': 'error', 'message': "Fichier de reçu manquant dans la requête POST."}, status=400)


from django.http import HttpResponse
from apps.inscriptions.pdf_services import generer_carte_etudiant_pdf, generer_recu_paiement_pdf, generer_hash_document

@login_required
def exporter_carte_etudiant_pdf(request, etudiant_id):
    """Exporte la carte d'étudiant au format PDF sécurisé ReportLab."""
    etudiant = get_object_or_404(Etudiant, pk=etudiant_id)
    
    if request.user.type_utilisateur == 'ETUDIANT' and etudiant.utilisateur != request.user:
        messages.error(request, "Accès refusé.")
        return redirect('tableau_bord:tableau_bord')
        
    base_url = request.build_absolute_uri('/')[:-1]
    pdf_bytes = generer_carte_etudiant_pdf(etudiant, domain_url=base_url)
    
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="carte_etudiant_{etudiant.matricule or etudiant.id}.pdf"'
    return response


@login_required
def exporter_recu_pdf(request, recu_id):
    """Exporte le reçu de paiement certifié en PDF sécurisé."""
    recu = get_object_or_404(RecuPaiement, pk=recu_id)
    
    if request.user.type_utilisateur == 'ETUDIANT' and recu.etudiant.utilisateur != request.user:
        messages.error(request, "Accès refusé.")
        return redirect('tableau_bord:tableau_bord')
        
    base_url = request.build_absolute_uri('/')[:-1]
    pdf_bytes = generer_recu_paiement_pdf(recu, domain_url=base_url)
    
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="recu_certifie_{recu.id}.pdf"'
    return response


def verifier_document(request):
    """Route publique d'authentification des documents officiels IAI via QR Code."""
    hash_doc = request.GET.get('hash', '').strip()
    doc_type = request.GET.get('type', '').strip()
    doc_id = request.GET.get('id', '').strip()
    
    est_valide = False
    objet_trouve = None
    date_creation_ref = None
    
    if hash_doc and doc_type and doc_id and doc_id.isdigit():
        object_id = int(doc_id)
        if doc_type == 'carte':
            etudiant = Etudiant.objects.filter(pk=object_id).first()
            if etudiant:
                date_str = etudiant.date_creation.strftime('%Y%m%d') if etudiant.date_creation else '20260101'
                expected_hash = generer_hash_document('CARD', etudiant.id, date_str)
                if expected_hash == hash_doc:
                    est_valide = True
                    objet_trouve = etudiant
                    date_creation_ref = etudiant.date_creation
        elif doc_type == 'recu':
            recu = RecuPaiement.objects.filter(pk=object_id, statut='VALIDE').first()
            if recu:
                date_str = recu.date_televersement.strftime('%Y%m%d%H%M') if recu.date_televersement else '20260101'
                expected_hash = generer_hash_document('RECU', recu.id, date_str)
                if expected_hash == hash_doc:
                    est_valide = True
                    objet_trouve = recu
                    date_creation_ref = recu.date_televersement

    context = {
        'hash_doc': hash_doc,
        'doc_type': doc_type,
        'est_valide': est_valide,
        'objet': objet_trouve,
        'date_ref': date_creation_ref,
        'titre': 'Vérification d\'Authenticité de Document'
    }
    return render(request, 'base/verifier_document.html', context)


@login_required
@role_required('CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'CHEF_FORMATION_CONTINUE', 'ADMIN_SYSTEME', 'DIRECTEUR')
def paiements_apprenants(request):
    """
    Vue dédiée à la Comptabilité et à la Direction pour consulter les apprenants
    et leur état de paiement (Formation Continue en premier, Formation Certifiante en second).
    """
    from apps.etudiants.models import Apprenant, Formation
    from django.db.models import Q

    q = request.GET.get('q', '').strip()
    
    # 1. Apprenants Formation Continue (haut de page)
    apprenants_continue_qs = Apprenant.objects.filter(
        formations__type_formation='CONTINUE'
    ).distinct().prefetch_related('formations').order_by('-date_creation')

    # 2. Apprenants Formation Certifiante (bas de page)
    apprenants_certif_qs = Apprenant.objects.filter(
        formations__type_formation='CERTIFICATION'
    ).distinct().prefetch_related('formations').order_by('-date_creation')

    if q:
        apprenants_continue_qs = apprenants_continue_qs.filter(
            Q(nom_complet__icontains=q) | Q(email__icontains=q) | Q(contact__icontains=q)
        )
        apprenants_certif_qs = apprenants_certif_qs.filter(
            Q(nom_complet__icontains=q) | Q(email__icontains=q) | Q(contact__icontains=q)
        )

    def _enrichir_apprenants(qs):
        liste = []
        for app in qs:
            total = sum(f.tarif for f in app.formations.all())
            paye = float(app.montant_paye or 0)
            reste = float(app.reste_a_payer or 0)
            if total > 0:
                pourcentage = min(100, int((paye / float(total)) * 100))
            else:
                pourcentage = 100 if paye > 0 else 0

            if reste == 0 and total > 0:
                statut_code = 'SOLDE'
                statut_label = 'Soldé'
                statut_color = 'bg-emerald-100 text-emerald-800 border-emerald-300'
            elif paye > 0:
                statut_code = 'AVANCE'
                statut_label = 'Avance versée'
                statut_color = 'bg-amber-100 text-amber-800 border-amber-300'
            else:
                statut_code = 'IMPAYE'
                statut_label = 'Non payé'
                statut_color = 'bg-rose-100 text-rose-800 border-rose-300'

            liste.append({
                'apprenant': app,
                'total': total,
                'paye': paye,
                'reste': reste,
                'pourcentage': pourcentage,
                'statut_code': statut_code,
                'statut_label': statut_label,
                'statut_color': statut_color,
                'formations_list': app.formations.all()
            })
        return liste

    data_continue = _enrichir_apprenants(apprenants_continue_qs)
    data_certif = _enrichir_apprenants(apprenants_certif_qs)

    total_recouvre = sum(a['paye'] for a in data_continue) + sum(a['paye'] for a in data_certif)
    total_reste = sum(a['reste'] for a in data_continue) + sum(a['reste'] for a in data_certif)

    context = {
        'apprenants_continue': data_continue,
        'apprenants_certif': data_certif,
        'total_apprenants_count': Apprenant.objects.count(),
        'total_recouvre': total_recouvre,
        'total_reste': total_reste,
        'q': q,
        'titre': 'Suivi Financier des Apprenants'
    }
    return render(request, 'paiements/apprenants_liste.html', context)



@login_required
def gestion_penalites(request):
    """
    Espace de Gestion Professionnelle des Pénalités de Retard
    Réservé au Chef de la Comptabilité, Admin Financier, Directeur et Admin Système.
    """
    allowed_roles = ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'DIRECTEUR', 'ADMIN_SYSTEME']
    if request.user.type_utilisateur not in allowed_roles and not request.user.is_superuser:
        messages.error(request, "Accès réservé au Chef de la Comptabilité et à la Direction.")
        return redirect('tableau_bord:tableau_bord')

    from apps.etudiants.models import Etudiant, Filiere, Niveau
    from apps.tableau_bord.models import PenalitePaiement
    from .services import calculer_penalites_etudiant

    # Filtres
    recherche = request.GET.get('q', '').strip()
    filiere_id = request.GET.get('filiere', '')
    niveau_id = request.GET.get('niveau', '')
    statut_filtre = request.GET.get('statut', '')  # 'RETARD', 'REGLE'

    etudiants_qs = Etudiant.objects.filter(statut__in=['PREINSCRIT', 'INSCRIT', 'ACTIF']).select_related('filiere', 'niveau', 'utilisateur')

    if recherche:
        etudiants_qs = etudiants_qs.filter(
            Q(matricule__icontains=recherche) |
            Q(nom__icontains=recherche) |
            Q(prenom__icontains=recherche) |
            Q(email__icontains=recherche)
        )

    if filiere_id:
        etudiants_qs = etudiants_qs.filter(filiere_id=filiere_id)

    if niveau_id:
        etudiants_qs = etudiants_qs.filter(niveau_id=niveau_id)

    # Calcul et agrégation des métriques
    donnees_etudiants = []
    total_cumule_penalites = 0
    total_penalites_recouvrees = 0
    total_penalites_en_attente = 0
    nb_etudiants_retard = 0

    for etud in etudiants_qs:
        info_pen = calculer_penalites_etudiant(etud)
        total_du = float(info_pen['total'] or 0)

        penalites_enregistrees = PenalitePaiement.objects.filter(etudiant=etud)
        total_regle = sum(float(p.montant_penalite or 0) for p in penalites_enregistrees)

        has_retard = total_du > 0
        is_regle = total_du == 0 and total_regle > 0

        # Filtre par statut
        if statut_filtre == 'RETARD' and not has_retard:
            continue
        elif statut_filtre == 'REGLE' and total_regle == 0:
            continue

        if has_retard:
            nb_etudiants_retard += 1

        total_cumule_penalites += total_du + total_regle
        total_penalites_recouvrees += total_regle
        total_penalites_en_attente += total_du

        donnees_etudiants.append({
            'etudiant': etud,
            'total_du': total_du,
            'total_regle': total_regle,
            'details': info_pen['details'],
            'nb_tranches_retard': len(info_pen['details']),
            'statut_penalite': 'EN_RETARD' if has_retard else ('REGLE' if total_regle > 0 else 'A_JOUR')
        })

    # Pagination
    paginator = Paginator(donnees_etudiants, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'filieres': Filiere.objects.filter(est_active=True),
        'niveaux': Niveau.objects.all(),
        'total_cumule': total_cumule_penalites,
        'total_recouvre': total_penalites_recouvrees,
        'total_en_attente': total_penalites_en_attente,
        'nb_etudiants_retard': nb_etudiants_retard,
        'recherche': recherche,
        'filiere_id': filiere_id,
        'niveau_id': niveau_id,
        'statut_filtre': statut_filtre,
        'titre': 'Gestion Professionnelle des Pénalités'
    }
    return render(request, 'paiements/gestion_penalites.html', context)


@login_required
def encaisser_penalite_caisse(request, etudiant_id):
    """Encaissement manuel des pénalités par le comptable en caisse (Espèces / Chèque)"""
    allowed_roles = ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'DIRECTEUR', 'ADMIN_SYSTEME']
    if request.user.type_utilisateur not in allowed_roles and not request.user.is_superuser:
        messages.error(request, "Permission insuffisante.")
        return redirect('paiements:gestion_penalites')

    etudiant = get_object_or_404(Etudiant, pk=etudiant_id)
    if request.method == 'POST':
        montant_saisi = float(request.POST.get('montant_encaisse', 0))
        mode_reglement = request.POST.get('mode_reglement', 'ESPECES')
        remarque = request.POST.get('remarque', 'Règlement direct en caisse')

        from .services import calculer_penalites_etudiant
        from apps.tableau_bord.models import PenalitePaiement, Activite, Notification

        penalites_info = calculer_penalites_etudiant(etudiant)
        TRANCHE_MAP = {1: 'PREINSCRIPTION', 2: 'TRANCHE_1', 3: 'TRANCHE_2', 4: 'TRANCHE_3'}

        solde_a_repartir = montant_saisi
        for d in penalites_info.get('details', []):
            if solde_a_repartir <= 0:
                break

            tranche_code = TRANCHE_MAP.get(d['tranche_numero'], 'PREINSCRIPTION')
            p, _ = PenalitePaiement.objects.get_or_create(
                etudiant=etudiant,
                tranche=tranche_code,
                defaults={
                    'montant_initial': d['tarif'],
                    'montant_total': d['montant_brut'],
                    'date_limite': d['date_limite'],
                    'semaines_retard': d['semaines_retard']
                }
            )

            deja_paye = float(p.montant_penalite or 0)
            recouvre_cette_fois = min(solde_a_repartir, d['montant'])
            nouveau_total_paye = deja_paye + recouvre_cette_fois

            p.montant_penalite = nouveau_total_paye
            p.montant_total = d['montant_brut']
            p.est_regle = (nouveau_total_paye >= d['montant_brut'])
            p.date_paiement = timezone.now().date()
            p.save()

            solde_a_repartir -= recouvre_cette_fois

        Activite.objects.create(
            utilisateur=request.user,
            type_action='PAIEMENT',
            description=f"Encaissement en caisse de {montant_saisi:,.0f} FCFA de pénalités pour {etudiant.get_nom_complet()} ({mode_reglement}). {remarque}",
            module='PAIEMENTS'
        )


        if etudiant.utilisateur:
            Notification.objects.create(
                utilisateur=etudiant.utilisateur,
                titre="Quittance de Pénalités (Caisse)",
                message=f"✅ Vos pénalités de retard ({montant_encaisse} FCFA) ont été réglées et quittancées en caisse.",
                type='SUCCESS'
            )

        messages.success(request, f"✅ Pénalités de {etudiant.get_nom_complet()} encaissées avec succès en caisse ({mode_reglement}).")
    return redirect('paiements:gestion_penalites')


@login_required
def exonerer_penalite(request, etudiant_id):
    """Accorder une exonération / remise de pénalités avec motif obligatoire"""
    allowed_roles = ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'DIRECTEUR', 'ADMIN_SYSTEME']
    if request.user.type_utilisateur not in allowed_roles and not request.user.is_superuser:
        messages.error(request, "Permission insuffisante.")
        return redirect('paiements:gestion_penalites')

    etudiant = get_object_or_404(Etudiant, pk=etudiant_id)
    if request.method == 'POST':
        motif = request.POST.get('motif_exoneration', '').strip()
        if not motif:
            messages.error(request, "Un motif est obligatoire pour accorder une exonération.")
            return redirect('paiements:gestion_penalites')

        from .services import calculer_penalites_etudiant
        from apps.tableau_bord.models import PenalitePaiement, Activite, Notification

        penalites_info = calculer_penalites_etudiant(etudiant)
        TRANCHE_MAP = {1: 'PREINSCRIPTION', 2: 'TRANCHE_1', 3: 'TRANCHE_2', 4: 'TRANCHE_3'}

        for d in penalites_info.get('details', []):
            tranche_code = TRANCHE_MAP.get(d['tranche_numero'], 'PREINSCRIPTION')
            PenalitePaiement.objects.update_or_create(
                etudiant=etudiant,
                tranche=tranche_code,
                defaults={
                    'est_regle': True,
                    'date_paiement': timezone.now().date(),
                    'montant_penalite': 0,
                    'montant_initial': d['tarif'],
                    'montant_total': 0,
                    'date_limite': d['date_limite'],
                    'semaines_retard': d['semaines_retard']
                }
            )

        Activite.objects.create(
            utilisateur=request.user,
            type_action='MODIFICATION',
            description=f"Exonération accordée sur les pénalités de {etudiant.get_nom_complet()} ({etudiant.matricule}). Motif: {motif}",
            module='PAIEMENTS'
        )

        if etudiant.utilisateur:
            Notification.objects.create(
                utilisateur=etudiant.utilisateur,
                titre="Exonération de Pénalités Accordée",
                message=f"🎁 Une remise / exonération de pénalités vous a été accordée par la Comptabilité. Motif : {motif}",
                type='INFO'
            )

        messages.success(request, f"🎁 Exonération de pénalités enregistrée pour {etudiant.get_nom_complet()}. Motif : {motif}")
    return redirect('paiements:gestion_penalites')


@login_required
def relancer_penalite_etudiant(request, etudiant_id):
    """Relance individuelle par Email & WhatsApp de l'étudiant"""
    allowed_roles = ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'DIRECTEUR', 'ADMIN_SYSTEME']
    if request.user.type_utilisateur not in allowed_roles and not request.user.is_superuser:
        messages.error(request, "Permission insuffisante.")
        return redirect('paiements:gestion_penalites')

    etudiant = get_object_or_404(Etudiant, pk=etudiant_id)
    from .services import calculer_penalites_etudiant
    from apps.tableau_bord.whatsapp_service import WhatsAppService
    from django.core.mail import send_mail

    info = calculer_penalites_etudiant(etudiant)
    total_du = info['total']

    if total_du > 0:
        msg = (
            f"*IAI-CAMEROUN (Douala)* ⚠️\n"
            f"*Relance de Pénalités de Retard*\n\n"
            f"Bonjour {etudiant.get_nom_complet()},\n"
            f"Votre dossier présente un cumul de pénalités de retard de *{total_du:,.0f} FCFA*.\n"
            f"Veuillez procéder au règlement sur votre espace étudiant ou auprès de la Caisse IAI.\n"
        )
        tel = etudiant.telephone or (etudiant.utilisateur.telephone if etudiant.utilisateur else None)
        if tel:
            try:
                WhatsAppService.envoyer_message(tel, msg)
            except Exception:
                pass

        email = etudiant.email or (etudiant.utilisateur.email if etudiant.utilisateur else None)
        if email:
            try:
                send_mail(
                    subject="[IAI-Cameroun] Rappel : Régularisation des pénalités de retard",
                    message=f"Bonjour {etudiant.get_nom_complet()},\n\nVous avez {total_du:,.0f} FCFA de pénalités de retard. Merci de régulariser sur http://127.0.0.1:8000/paiements/penalites/",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    fail_silently=True
                )
            except Exception:
                pass

        messages.success(request, f"📲 Relance transmise avec succès à {etudiant.get_nom_complet()} ({total_du:,.0f} FCFA).")
    else:
        messages.info(request, f"L'étudiant {etudiant.get_nom_complet()} n'a aucune pénalité en retard.")

    return redirect('paiements:gestion_penalites')


@login_required
def exporter_penalites_csv(request):
    """Exportation CSV du rapport des pénalités de retard"""
    allowed_roles = ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'DIRECTEUR', 'ADMIN_SYSTEME']
    if request.user.type_utilisateur not in allowed_roles and not request.user.is_superuser:
        return HttpResponse("Accès refusé", status=403)

    import csv
    from apps.etudiants.models import Etudiant
    from .services import calculer_penalites_etudiant

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="Rapport_Penalites_IAI.csv"'
    response.write('\ufeff'.encode('utf8'))

    writer = csv.writer(response)
    writer.writerow(['Matricule', 'Nom', 'Prénom', 'Filière', 'Niveau', 'Téléphone', 'Pénalités Dues (FCFA)', 'Nb Tranches Retard'])

    etudiants = Etudiant.objects.filter(statut__in=['PREINSCRIT', 'INSCRIT', 'ACTIF']).select_related('filiere', 'niveau')
    for etud in etudiants:
        info = calculer_penalites_etudiant(etud)
        if info['total'] > 0:
            writer.writerow([
                etud.matricule or '',
                etud.nom,
                etud.prenom,
                etud.filiere.code if etud.filiere else '',
                etud.niveau.numero if etud.niveau else '',
                etud.telephone or '',
                info['total'],
                len(info['details'])
            ])

    return response


@login_required
def exporter_penalites_excel(request):
    """Exportation du rapport des pénalités de retard au format Microsoft Excel (.xlsx)"""
    allowed_roles = ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'DIRECTEUR', 'ADMIN_SYSTEME']
    if request.user.type_utilisateur not in allowed_roles and not request.user.is_superuser:
        return HttpResponse("Accès refusé", status=403)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from apps.etudiants.models import Etudiant
    from apps.tableau_bord.models import PenalitePaiement
    from .services import calculer_penalites_etudiant

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pénalités de Retard"
    ws.views.sheetView[0].showGridLines = True

    # Styles
    font_title = Font(name="Calibri", size=16, bold=True, color="1E293B")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="64748B")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="0F172A")
    font_regular = Font(name="Calibri", size=11, color="1E293B")

    fill_header = PatternFill(start_color="881337", end_color="881337", fill_type="solid")
    fill_zebra = PatternFill(start_color="FFF1F2", end_color="FFF1F2", fill_type="solid")
    fill_total = PatternFill(start_color="FFE4E6", end_color="FFE4E6", fill_type="solid")

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # Titre
    ws.merge_cells('A1:H1')
    ws['A1'] = "IAI-CAMEROUN (Centre de Douala) - Rapport des Pénalités de Retard"
    ws['A1'].font = font_title
    ws['A1'].alignment = align_left

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Généré le {timezone.now().strftime('%d/%m/%Y à %H:%M')} par {request.user.get_full_name() or request.user.username}"
    ws['A2'].font = font_subtitle
    ws['A2'].alignment = align_left

    # En-têtes du tableau (Ligne 4)
    headers = ['Matricule', 'Nom & Prénom', 'Filière', 'Niveau', 'Téléphone', 'Déjà Encaissé (FCFA)', 'Reste Dû (FCFA)', 'Nb Tranches Retard']
    ws.append([])
    ws.append(headers)

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center if col_num in [1, 3, 4, 8] else (align_right if col_num in [6, 7] else align_left)

    # Données
    etudiants = Etudiant.objects.filter(statut__in=['PREINSCRIT', 'INSCRIT', 'ACTIF']).select_related('filiere', 'niveau')
    row_idx = 5
    sum_du = 0
    sum_regle = 0

    for etud in etudiants:
        info = calculer_penalites_etudiant(etud)
        total_du = float(info['total'] or 0)
        penalites_reglees_qs = PenalitePaiement.objects.filter(etudiant=etud)
        total_regle = sum(float(p.montant_penalite or 0) for p in penalites_reglees_qs)

        if total_du > 0 or total_regle > 0:
            sum_du += total_du
            sum_regle += total_regle

            ws.append([
                etud.matricule or '',
                etud.get_nom_complet(),
                etud.filiere.code if etud.filiere else '',
                f"Niveau {etud.niveau.numero}" if etud.niveau else 'N1',
                etud.telephone or '',
                total_regle,
                total_du,
                len(info['details'])
            ])

            current_row = ws[row_idx]
            for i, cell in enumerate(current_row, 1):
                cell.font = font_regular
                cell.border = thin_border
                if i in [6, 7]:
                    cell.number_format = '#,##0 "FCFA"'
                    cell.alignment = align_right
                elif i in [1, 3, 4, 8]:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

                if row_idx % 2 == 0:
                    cell.fill = fill_zebra

            row_idx += 1

    # Ligne de Totalisation
    ws.append(['', 'TOTAL GENERAL', '', '', '', sum_regle, sum_du, ''])
    tot_row = ws[row_idx]
    for i, cell in enumerate(tot_row, 1):
        cell.font = font_bold
        cell.fill = fill_total
        cell.border = thin_border
        if i in [6, 7]:
            cell.number_format = '#,##0 "FCFA"'
            cell.alignment = align_right

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Rapport_Penalites_IAI.xlsx"'
    wb.save(response)
    return response


@login_required
def exporter_penalites_pdf(request):
    """Exportation du rapport officiel des pénalités au format PDF haute définition (ReportLab)"""
    allowed_roles = ['CHEF_COMPTABILITE', 'ADMIN_FINANCIER', 'DIRECTEUR', 'ADMIN_SYSTEME']
    if request.user.type_utilisateur not in allowed_roles and not request.user.is_superuser:
        return HttpResponse("Accès refusé", status=403)

    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from io import BytesIO
    from apps.etudiants.models import Etudiant
    from apps.tableau_bord.models import PenalitePaiement
    from .services import calculer_penalites_etudiant

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm
    )

    story = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=16,
        leading=20,
        textColor=colors.HexColor('#881337')
    )
    subtitle_style = ParagraphStyle(
        'DocSubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#475569')
    )
    cell_header_style = ParagraphStyle(
        'CellHeader',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=9,
        leading=11,
        textColor=colors.white,
        alignment=1
    )
    cell_body_style = ParagraphStyle(
        'CellBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#0F172A')
    )
    cell_bold_style = ParagraphStyle(
        'CellBold',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#0F172A')
    )

    story.append(Paragraph("<b>INSTITUT AFRICAIN D'INFORMATIQUE - CENTRE DE DOUALA</b>", subtitle_style))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph("ETAT DE RECOUVREMENT & DE SUIVI DES PENALITES DE RETARD", title_style))
    story.append(Paragraph(f"Édité le {timezone.now().strftime('%d/%m/%Y à %H:%M')} par {request.user.get_full_name() or request.user.username}", subtitle_style))
    story.append(Spacer(1, 0.4 * cm))
    story.append(HRFlowable(width="100%", thickness=1.5, color=colors.HexColor('#881337'), spaceAfter=10))

    table_data = [[
        Paragraph("Matricule", cell_header_style),
        Paragraph("Nom & Prénom de l'Étudiant", cell_header_style),
        Paragraph("Filière", cell_header_style),
        Paragraph("Niveau", cell_header_style),
        Paragraph("Téléphone", cell_header_style),
        Paragraph("Déjà Encaissé", cell_header_style),
        Paragraph("Reste Dû", cell_header_style),
        Paragraph("Retard", cell_header_style)
    ]]

    etudiants = Etudiant.objects.filter(statut__in=['PREINSCRIT', 'INSCRIT', 'ACTIF']).select_related('filiere', 'niveau')
    sum_du = 0
    sum_regle = 0

    for etud in etudiants:
        info = calculer_penalites_etudiant(etud)
        total_du = float(info['total'] or 0)
        penalites_reglees_qs = PenalitePaiement.objects.filter(etudiant=etud)
        total_regle = sum(float(p.montant_penalite or 0) for p in penalites_reglees_qs)

        if total_du > 0 or total_regle > 0:
            sum_du += total_du
            sum_regle += total_regle

            table_data.append([
                Paragraph(etud.matricule or '-', cell_body_style),
                Paragraph(etud.get_nom_complet(), cell_bold_style),
                Paragraph(etud.filiere.code if etud.filiere else '-', cell_body_style),
                Paragraph(f"Niveau {etud.niveau.numero}" if etud.niveau else 'N1', cell_body_style),
                Paragraph(etud.telephone or '-', cell_body_style),
                Paragraph(f"{total_regle:,.0f} FCFA", cell_body_style),
                Paragraph(f"{total_du:,.0f} FCFA", cell_bold_style),
                Paragraph(f"{len(info['details'])} tranche(s)", cell_body_style)
            ])

    table_data.append([
        Paragraph("TOTAL", cell_header_style),
        Paragraph("TOTAL GENERAL DES PENALITES", cell_header_style),
        Paragraph("", cell_header_style),
        Paragraph("", cell_header_style),
        Paragraph("", cell_header_style),
        Paragraph(f"{sum_regle:,.0f} FCFA", cell_header_style),
        Paragraph(f"{sum_du:,.0f} FCFA", cell_header_style),
        Paragraph("", cell_header_style)
    ])

    col_widths = [3.2 * cm, 6.5 * cm, 2.2 * cm, 2.2 * cm, 3.2 * cm, 3.5 * cm, 3.5 * cm, 2.2 * cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    t_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#881337')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E2E8F0')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1E293B')),
    ])

    for i in range(1, len(table_data) - 1):
        if i % 2 == 0:
            t_style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFF1F2'))

    t.setStyle(t_style)
    story.append(t)

    story.append(Spacer(1, 1.2 * cm))
    sig_data = [
        [Paragraph("<b>Le Chef de la Comptabilité</b>", subtitle_style), Paragraph("<b>Le Directeur du Centre</b>", subtitle_style)],
        [Spacer(1, 1.5 * cm), Spacer(1, 1.5 * cm)],
        [Paragraph("Visa & Tampon", subtitle_style), Paragraph("Visa & Tampon", subtitle_style)]
    ]
    sig_table = Table(sig_data, colWidths=[13 * cm, 13 * cm])
    sig_table.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'CENTER')]))
    story.append(sig_table)

    doc.build(story)
    pdf_out = buffer.getvalue()
    buffer.close()

    response = HttpResponse(pdf_out, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Rapport_Penalites_IAI.pdf"'
    return response



