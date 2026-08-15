"""
Helper d'export Excel/CSV enrichi avec l'en-tête officiel IAI-Cameroun
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_TITLE_INSTITUTION = "INSTITUT AFRICAIN D'INFORMATIQUE - CENTRE D'EXCELLENCE TECHNOLOGIQUE PAUL BIYA"
HEADER_SUBTITLE = "Représentation du Cameroun — BP 13 719 Yaoundé | Tél. (237) 242 72 99 57 / 242 72 99 58 | contact@iaicameroun.com"

def generer_excel_officiel(titre_document, colonnes, donnees, nom_feuille="Données"):
    """
    Génère un classeur Excel avec en-tête institutionnel IAI-Cameroun et mise en forme professionnelle.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nom_feuille

    # Style des en-têtes
    font_inst = Font(name='Arial', size=12, bold=True, color='004D25')
    font_sub = Font(name='Arial', size=9, italic=True, color='555555')
    font_title = Font(name='Arial', size=14, bold=True, color='000000')
    font_th = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    
    fill_th = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')

    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )

    num_cols = len(colonnes)
    max_col_letter = get_column_letter(max(num_cols, 1))

    # Ligne 1 : Nom de l'établissement
    ws.merge_cells(f'A1:{max_col_letter}1')
    ws['A1'] = HEADER_TITLE_INSTITUTION
    ws['A1'].font = font_inst
    ws['A1'].alignment = align_center

    # Ligne 2 : Sous-titre / Contacts
    ws.merge_cells(f'A2:{max_col_letter}2')
    ws['A2'] = HEADER_SUBTITLE
    ws['A2'].font = font_sub
    ws['A2'].alignment = align_center

    # Ligne 4 : Titre du document
    ws.merge_cells(f'A4:{max_col_letter}4')
    ws['A4'] = titre_document.upper()
    ws['A4'].font = font_title
    ws['A4'].alignment = align_center

    # Ligne 6 : En-têtes des colonnes
    start_row = 6
    for col_idx, col_name in enumerate(colonnes, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.font = font_th
        cell.fill = fill_th
        cell.alignment = align_center
        cell.border = thin_border

    # Remplissage des données
    for row_idx, row_data in enumerate(donnees, start_row + 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = align_left
            cell.border = thin_border

    # Ajustement automatique de la largeur des colonnes
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
